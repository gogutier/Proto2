from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q
from django.contrib.auth.forms import UserCreationForm
from django.views.generic.edit import CreateView
from django.forms.models import model_to_dict
from django.utils import timezone
from blog.models import Post,Comment, appointment, CargaCSV, OCImportacion, ProdID, Book, PruebaMod, PruebaTabla, OrdenProg, DetalleProg, ProdReal, Maquinas, Turnos, Minuta, OrderInfo, Padron, DiaConv2, OrdenProgCorr, DetalleProgCorr, Meses, Semanas, FotoInventario, ProyMkt, ProyMktMes, ProyMktPadron, ProdRealCorr, InfoWIP, Camion, OrdenCorrplan, FotoCorrplan, Cartones, CalleBPT, BobInvCic, MovPallets, Pallet, UbicPallet, PalletCic, TomaInvCic, DatosWIP_Prog, Datos_Proy_WIP, IDProgCorr, Datos_MovPallets, Datos_MovPallets_B, Foto_Datos_MovPallets, Datos_Inv_WIP, Foto_Datos_Inv_WIP, FiltroEntradaWIP, FiltroSalidaWIP, Foto_Inv_Cic_WIP, Foto_Calles_Inv_Cic_WIP, Foto_Palletscti_Inv_Cic_WIP, Foto_Palletsencontrados_Inv_Cic_WIP, Foto_Palletsenotracalle_Inv_Cic_WIP, Foto_Palletsnoencontrados_Inv_Cic_WIP, MovRollos, ConsumoRollos, Foto_ConsumoRollos, m2Maqruta_WIP
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from blog.forms import PostForm, CommentForm, ContactForm, AppointmentForm, OCImportacionForm, ProdIDForm, BookFormset, BookModelFormset, PruebaModForm, MinutaForm, QRForm, MovPalletForm
from django.views.generic import (TemplateView,ListView,CreateView,DetailView, UpdateView, DeleteView, View)
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import csv
from datetime import datetime, timedelta
from io import StringIO
import pruebawebscrap
import webscrap2
import openpyxl
from django.http.response import HttpResponse

import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
from excel_utils import WriteToExcel
from excel_utils import WriteToExcel2
import pruebawebscrapdowntime

#VIEWS ES DONDE SE PUEDE PROGRAMR EN PYTHON?
#views functions take as input: HTTPRESPONSE objects, and returns HTTPRESpose object (html output)


@csrf_exempt
def get_data_busqueda_improductivos(request, *args, **kwargs):

        #consulto en la base de datos todos los objetos pallet que tiene ubicación zTCY1 (a minúsculas). sumo sus m2 por pallets. los Cuento


        fechainicio=datetime.strptime(request.POST['fechainicio'], "%Y-%m-%d")
        fechafin=datetime.strptime(request.POST['fechafin'], "%Y-%m-%d")
        maquina=int(request.POST['maquina'])
        prueba={"prueba1":(33,323), "prueba2":{"A":3,"B":4}}


        if fechafin-fechainicio<=timedelta(days=31):
            resultado= pruebawebscrapdowntime.webscrap_wip(maquina, fechainicio, fechafin)

            maq=""
            if maquina==1:
                maq="Corrugado"
            if maquina==2:
                maq="FFG"
            if maquina==5:
                maq="DRO"
            if maquina==27:
                maq="FFW"
            if maquina==12:
                maq="WRD"
            if maquina==11:
                maq="HCR"
            if maquina==4:
                maq="TCY"

            mensaje="Tiempos improducivos para "+ maq
        else:
            mensaje="Rango máximo de 31 días!!!"
            resultado= [[0,0,0,"Reducir rango búsqueda",0,0,0]]
        #print(resultado)



        data = {
        "resultado":list(resultado),
        "mensaje":mensaje,


        }
        print("Enviando datos improductivos")
        return JsonResponse(data)#http response con el datatype de JS

def busqueda_improductivos(request):
    #print("cargando consumos puestos")
    template_name = 'blog/busqueda_improductivos.html'

    if (request.method == "POST"):
        if 'excel' in request.POST:
            print("Recibiendo respuesta excel en POST")
            fechainicio=datetime.strptime(request.POST['fechainicio'], "%Y-%m-%d")
            fechafin=datetime.strptime(request.POST['fechafin'], "%Y-%m-%d")
            maquina=int(request.POST['maquina'])


            resultado= pruebawebscrapdowntime.webscrap_wip(maquina, fechainicio, fechafin)

            maq=""
            if maquina==1:
                maq="Corrugado"
            if maquina==2:
                maq="FFG"
            if maquina==5:
                maq="DRO"
            if maquina==27:
                maq="FFW"
            if maquina==12:
                maq="WRD"
            if maquina==11:
                maq="HCR"
            if maquina==4:
                maq="TCY"


            datosinv = list(resultado)
            town = None
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
            xlsx_data = WriteToExcel2(datosinv, maq)
            response.write(xlsx_data)
            return response



    return render(request, template_name, {})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?








def panel_consumos_conv_efi(request):
    #print("cargando datos wip")
    template_name = 'blog/panel_consumos_conv_efi.html'

    print("Calculando consumos en base a movimientos EFI")

    #lista de movimientos de máquinas conv:

    listamaqs=("HCR","TCY","WRD", "DIM", "FFW", "DRO", "FFG")
    #filtroconvqs=Q()
    m2Ingresado={}
    totIngreso=0
    for ubic in listamaqs:
    #    filtroconvqs = filtroconvqs | Q(DESTINATION=ubic)
        consumos=MovPallets.objects.filter(DESTINATION=ubic, OPERATIONNO=0, EVENTDATETIME__gte=datetime.now().replace(day=1, hour= 7, minute=0, second=0, microsecond=0), EVENTDATETIME__lt=datetime.now() )
        m2Ingresado[ubic]=0
        for consumo in consumos:
            m2Ingresado[ubic]+=consumo.m2pallet
        m2Ingresado[ubic]=round(m2Ingresado[ubic]/1000,1)
        totIngreso+=m2Ingresado[ubic]
    totIngreso=round(totIngreso,1)

    listacalles=("ZPICADO","ZPNC","ZFFG1","ZFFG2","ZDRO1","ZDRO2","ZFFW1","ZFFW2","ZSOB1","ZWRD1","ZWRD2","ZSOB2","ZHCR1","ZHCR2","ZTCY1","ZTCY2","ZPASILLO")
    filtrocalleqs=Q()
    for calle in listacalles:
        filtrocalleqs = filtrocalleqs | Q(DESTINATION=calle)
    #filtroconvqs=Q()

    m2Devuelto={}
    totDevuelto=0
    for ubic in listamaqs:
    #    filtroconvqs = filtroconvqs | Q(DESTINATION=ubic)
        devs=MovPallets.objects.filter(filtrocalleqs, SOURCE=ubic, OPERATIONNO=0, EVENTDATETIME__gte=datetime.now().replace(day=1, hour= 7, minute=0, second=0, microsecond=0), EVENTDATETIME__lt=datetime.now() )
        m2Devuelto[ubic]=0
        for dev in devs:
            m2Devuelto[ubic]+=dev.m2pallet
        m2Devuelto[ubic]=round(m2Devuelto[ubic]/1000,1)
        totDevuelto+=m2Devuelto[ubic]
    totDevuelto=round(totDevuelto,1)

    m2Apicado={}
    totApicado=0

    for ubic in listamaqs:
    #    filtroconvqs = filtroconvqs | Q(DESTINATION=ubic)
        devs=MovPallets.objects.filter(DESTINATION='ZPICADO', SOURCE=ubic, OPERATIONNO=0, EVENTDATETIME__gte=datetime.now().replace(day=1, hour= 7, minute=0, second=0, microsecond=0), EVENTDATETIME__lt=datetime.now() )
        m2Apicado[ubic]=0
        for dev in devs:
            m2Apicado[ubic]+=dev.m2pallet
        m2Apicado[ubic]=round(m2Apicado[ubic]/1000,1)
        totApicado+=m2Apicado[ubic]
    totApicado=round(totApicado,1)

    transportistas=[]

    m2ConsumoTotal={}
    totConsumoTotal=0
    for ubic in listamaqs:
        m2ConsumoTotal[ubic]=m2Ingresado[ubic]-m2Devuelto[ubic]
        m2ConsumoTotal[ubic] = round(m2ConsumoTotal[ubic],1)
        totConsumoTotal+=m2ConsumoTotal[ubic]
    totConsumoTotal=round(totConsumoTotal,1)
    print(m2ConsumoTotal)
    return render(request, template_name, {'consumos':consumos, 'transportistas':transportistas, 'm2Ingresado':m2Ingresado, 'm2Devuelto':m2Devuelto, 'm2Apicado':m2Apicado, 'm2ConsumoTotal':m2ConsumoTotal,'totConsumoTotal':totConsumoTotal, 'totApicado':totApicado, 'totDevuelto':totDevuelto, 'totIngreso':totIngreso  })#acá le puedo decir que los mande ordenados por fecha?



@csrf_exempt
def get_data_busqueda_pallet_wip(request, *args, **kwargs):

        #consulto en la base de datos todos los objetos pallet que tiene ubicación zTCY1 (a minúsculas). sumo sus m2 por pallets. los Cuento


        cliente=(request.POST['cliente'])
        padron=(request.POST['padron'])
        orderID=(request.POST['orderID'])
        prueba={"prueba1":(33,323), "prueba2":{"A":3,"B":4}}

        listacalles=("ZPICADO","ZPNC","ZFFG1","ZFFG2","ZDRO1","ZDRO2","ZFFW1","ZFFW2","ZSOB1","ZWRD1","ZWRD2","ZSOB2","ZHCR1","ZHCR2","ZTCY1","ZTCY2","CORR_UPPER_Stacker","CORR_UPPER_Stacker")
        filtrocalleqs=Q()
        for calle in listacalles:
            filtrocalleqs = filtrocalleqs | Q(ubic=calle)

        if orderID!="":
            resultado= Pallet.objects.filter(filtrocalleqs).filter(ORDERID=orderID).values()
        elif cliente!="":
            resultado= Pallet.objects.filter(filtrocalleqs).filter(cliente__icontains=cliente).order_by("-fechacreac")
            if padron!="":
                resultado=  resultado.filter(padron__icontains=padron).order_by("-fechacreac")




            resultado=resultado.values()

        data = {
        "resultado":list(resultado),


        }
        print("Enviando datos inventario")
        return JsonResponse(data)#http response con el datatype de JS

def busqueda_pallet_wip(request):
    #print("cargando consumos puestos")
    template_name = 'blog/busqueda_pallet_wip.html'

    return render(request, template_name, {})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?














def get_data_consumo_rollos(request, *args, **kwargs):

    labels={}
    ahora=datetime.now().replace(hour= 0, minute=0, second=0, microsecond=0)
    horizonte=35
    for i in range(0,horizonte+1):
        #por ahora los voy a ordenar por turno, después por hora.
        fechaini=(ahora-timedelta(days=horizonte-i) ).replace(hour= 7)
        fechafin=(ahora-timedelta(days=horizonte-i)).replace(hour= 14, minute=30)
        turno="A"
        label= fechaini.strftime("%d-%m") + " " + turno
        #calculo el m2 real convertido y corrugado en ese turno, para comparar con las salidas y entradas declaradas
        #print(m2Corr)
        labels.update({label:([],[],fechaini,fechafin)})#append({"fechaini":fechaini ,"fechafin":fechafin ,"turno":turno, "label": label})

        fechaini=(ahora-timedelta(days=horizonte-i)).replace(hour= 14, minute=30)
        fechafin=(ahora-timedelta(days=horizonte-i)).replace(hour= 22)
        turno="B"
        label= fechaini.strftime("%d-%m") + " " + turno
        #print(m2Corr)
        labels.update({label:([],[],fechaini,fechafin)})#labels.append({"fechaini":fechaini ,"fechafin":fechafin ,"turno":turno, "label": label})
        fechaini=(ahora-timedelta(days=horizonte-i)).replace(hour= 22)
        fechafin=(ahora-timedelta(days=horizonte-i-1)).replace(hour= 7)
        turno="C"
        label= fechaini.strftime("%d-%m") + " " + turno
        labels.update({label:([],[],fechaini,fechafin)})#labels.append({"fechaini":fechaini ,"fechafin":fechafin ,"turno":turno, "label": label})

        #Cada label corresponde a un turno, incluye una lista de movimientos y otra lista independiente de consumo de rollos.
        #print("ahora calculo las listas de lo que se declaró como ingreso y como salida al wip  ")

    print(labels)


    consumos=[]


    #Obtengo la foto de consumo de rollos más reciente:


    for label in labels:

        for mov in MovRollos.objects.filter(fecha__gte=labels[label][2], fecha__lt=labels[label][3]):

            print(mov)
            foto=Foto_ConsumoRollos.objects.filter(label=label).order_by('-fecha_foto')[0]

            #Acá actualmente estoy mostrando todos los rollos que se entregaron ese turno. Ahora sólo quiero mostrar los rollos que se entregaron pero que no tienen un consumo asociado en ese u otros turnos posteriores.
            #labels[label][1].append({"idrollo":mov.idrollo,"origen":mov.origen,"destino":mov.destino,"fecha":mov.fecha,"usuario":mov.usuario})
            movsinconsumo=ConsumoRollos.objects.filter(foto=foto, fechaini__gte=labels[label][2], RollID=mov.idrollo).count()
            if movsinconsumo==0:
                labels[label][1].append({"idrollo":mov.idrollo,"origen":mov.origen,"destino":mov.destino,"fecha":mov.fecha,"usuario":mov.usuario})




        try:
            print("consultando consumos para " + label)
            foto=Foto_ConsumoRollos.objects.filter(label=label).order_by('-fecha_foto')[0]


            for consumo in ConsumoRollos.objects.filter(foto=foto).order_by('grado'):
                print(consumo)
                #identifico el último movimiento de ese rollo al empalmador que corresponda:

            #    if entrega:
            #        labels[label][0].append({"fechaini":consumo.fechaini,"turno":consumo.turno,"RollID":consumo.RollID,"RollStandID":consumo.RollStandID,"formato":consumo.formato, "peso": consumo.peso, "grado":consumo.grado, "mlusados":consumo.mlusados, "mlrestantes":consumo.mlrestantes, "peelwaste":consumo.peelwaste, "fechaentrega":entrega.fecha})
            #    else:

                #labels[label][0][]['fechaentrega']=1
                dicc={"fechaini":consumo.fechaini,"turno":consumo.turno,"RollID":consumo.RollID,"RollStandID":consumo.RollStandID,"formato":consumo.formato, "peso": consumo.peso, "grado":consumo.grado, "mlusados":consumo.mlusados, "mlrestantes":consumo.mlrestantes, "peelwaste":consumo.peelwaste, "fechaentrega":0, "userentrega":"N/A" , "destinoentrega":"N/A"}

                try:
                    entrega=MovRollos.objects.filter(destinoaux=consumo.RollStandID[-2], idrollo=consumo.RollID, fecha__gte=labels[label][3]-timedelta(days=2), fecha__lt=labels[label][3]).order_by('-fecha')[0]

                    #print("destino Consumo:")
                    #print(consumo.RollStandID[-2])
                    #print("destinoMovPallet:")
                    #print(entrega.destino[-2])

                    dicc['fechaentrega']=entrega.fecha
                    dicc['userentrega']=entrega.usuario
                    dicc['destinoentrega']=entrega.destino
                    print("entrega encontrada!")

                except Exception as e:
                    print(e)
                    print("entrega no encontrada")

                print("consumo encontrado para " + label +": " + str(dicc))
                labels[label][0].append(dicc)


        except Exception as e:
            #print(e)
            print("Error al buscar consumo de rollo para fecha " + label)




    data = {
    "labels":labels,
    "consumos":consumos,
            }
    print("Enviando datos proy wip")
    return JsonResponse(data)#http response con el datatype de JS





def panel_consumo_rollos(request):
    #print("cargando consumos puestos")
    template_name = 'blog/panel_consumo_rollos.html'

    return render(request, template_name, {})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?




def carga_mov_rollos(request):
    pruebamods = PruebaMod.objects.all()

    template_name = 'blog/carga_mov_rollos.html'

    if request.method == "POST":
        form = PruebaModForm(request.POST)

        datocrudo=form.data["ultrafile"]

        datoprocesado=datocrudo.split(",;,")

        for i in range (len(datoprocesado)):
            datoprocesado[i]=datoprocesado[i].split(",")


            ####### identifica las columnas y crea los objetos que me interesanself.
        for dato in datoprocesado:

            print(dato[3])
            fecha=datetime.strptime(dato[3], "%d-%m-%Y  %H:%M:%S ")

            o, created=MovRollos.objects.get_or_create(idrollo=dato[0],origen=dato[1],destino=dato[2], fecha=fecha, usuario=dato[5])
            if(len(dato[2])>2):
                o.destinoaux=dato[2][-2]
            o.save()



            #for i in range(len(datoprocesado[0])):


    else:
        form = PruebaModForm()

    return render(request, template_name, {'form':form})



def panel_proy_wip(request):
    #print("cargando consumos puestos")
    template_name = 'blog/panel_proy_wip.html'

    return render(request, template_name, {})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?

def get_data_proy_wip(request, *args, **kwargs):

    labels=[]

    for dato in Datos_Proy_WIP.objects.all():

        labels.append({"fecha_inicio":dato.fecha_inicio,"fecha_fin":dato.fecha_fin,"turno":dato.turno, "label": dato.label, "M2Conv":dato.M2Conv, "M2Corr":dato.M2Corr, "M2Inv":dato.M2Inv, "M2ProyCorr":dato.M2ProyCorr})

    fotocorrplan = FotoCorrplan.objects.all()[0].fecha_foto.strftime("%m/%d/%Y %H:%M:%S")

    data = {
    "labels":labels,
    "fotocorrplan":fotocorrplan,
            }
    print("Enviando datos proy wip")
    return JsonResponse(data)#http response con el datatype de JS

def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename.file)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()

    for row in xrange(0, nrows):
        for col in xrange(0, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)

    return book1

def up_excel(request):
    if "GET" == request.method:
        return render(request, 'blog/up_excel.html', {})
    else :
        excel_file1 = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        excel_file2=open_xls_as_xlsx(excel_file1)

        wb = openpyxl.load_workbook(excel_file2)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Hoja1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'blog/up_excel.html', {"excel_data":excel_data})



def panel_movpallets(request):
    #print("cargando consumos puestos")
    template_name = 'blog/panel_movpallets.html'

    return render(request, template_name, {})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?

@csrf_exempt
def panel_inv_ciclico(request):
    #print("cargando consumos puestos")



    if request.method == "POST":

        if 'excel' not in request.POST:

            #Crea una nueva TomaInvCic.
            print("RECIBIENTO UN POST")
            o = TomaInvCic.objects.create()

            o.save()

            tomainvantiguos= TomaInvCic.objects.filter(fechatomainvcic__lt=o.fechatomainvcic)
            for toma in tomainvantiguos:
                toma.delete()

        #Borra Las anteriores.


    #LA parte para generar el excel_data
        if 'excel' in request.POST:
            print("Recibiendo respuesta excel en POST")

            #en datosinv es donde envío los datos a incluir en el gráfico
            #Obtengo la foto del inv cic más reciente



            ################## generando los mismos datos en el excel_data

            fotoinv=Foto_Inv_Cic_WIP.objects.all().order_by('-pk')[0]

            datosWIP2={}


            for calle in Foto_Calles_Inv_Cic_WIP.objects.filter(foto=fotoinv):

                lista2=[[],[]]
                datosWIP2[str(calle)]={}
                datosWIP2[str(calle)]['palletscti']=[[],[]]
                datosWIP2[str(calle)]['palletsencontrados']=[[],[]]
                datosWIP2[str(calle)]['palletsnoencontrados']=[[],[]]
                datosWIP2[str(calle)]['palletsenotracalle']=[[],[]]

                for pallet in Foto_Palletscti_Inv_Cic_WIP.objects.filter(calle=calle):
                    if pallet:
                        datosWIP2[str(calle)]['palletscti'][0].append(pallet.pallet)
                        datosWIP2[str(calle)]['palletscti'][1].append(pallet.ORDERID)
                        #print(pallet.pallet)
                        #print(pallet.ORDERID)

                for pallet in Foto_Palletsencontrados_Inv_Cic_WIP.objects.filter(calle=calle):
                    if pallet:
                        datosWIP2[str(calle)]['palletsencontrados'][0].append(pallet.pallet)
                        datosWIP2[str(calle)]['palletsencontrados'][1].append(pallet.ORDERID)
                        #print(pallet.pallet)
                        #print(pallet.ORDERID)

                for pallet in Foto_Palletsenotracalle_Inv_Cic_WIP.objects.filter(calle=calle):
                    if pallet:
                        datosWIP2[str(calle)]['palletsenotracalle'][0].append(pallet.pallet)
                        datosWIP2[str(calle)]['palletsenotracalle'][1].append(pallet.ORDERID)
                        #print(pallet.pallet)
                        #print(pallet.ORDERID)

                for pallet in Foto_Palletsnoencontrados_Inv_Cic_WIP.objects.filter(calle=calle):
                    if pallet:
                        datosWIP2[str(calle)]['palletsnoencontrados'][0].append(pallet.pallet)
                        datosWIP2[str(calle)]['palletsnoencontrados'][1].append(pallet.ORDERID)
                        #print(pallet.pallet)
                        #print(pallet.ORDERID)





                #####################






            datosinv = datosWIP2
            town = None
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
            xlsx_data = WriteToExcel(datosinv, town)
            response.write(xlsx_data)
            return response


    template_name = 'blog/panel_inv_ciclico.html'

    return render(request, template_name, {})#



def get_data_movpallets(request, *args, **kwargs):
    labels=[]
    labels2=[]
    #estructura: inicio turno en datetime, inicio turno en texto, inicio turno en descripción, m2 ingreso a planta en ese turno, m2 salida de planta en ese turno.

    #fecha, label, m2in, m2out

    #labels.append((ahora+timedelta(days=i)).replace(hour= 0, minute=0, second=0, microsecond=0))
    foto= Foto_Datos_MovPallets.objects.all().order_by('fecha_foto')[0]



    for dato in Datos_MovPallets.objects.filter(programa=foto):
        labels.append({"fecha":dato.fecha,"fechafin":dato.fechafin,"turno":dato.turno, "label": dato.label, "cantidadIn":dato.cantidadIn, "m2In":dato.m2In,"m2DeConvAPicado":dato.m2DeConvAPicado, "m2EntregadoAConv":dato.m2EntregadoAConv, "cantidadCorrPicado":dato.cantidadCorrPicado, "m2CorrPicado":dato.m2CorrPicado, "cantidadDirectoConv":dato.cantidadDirectoConv, "m2DirectoConv":dato.m2DirectoConv, "cantidadOut":dato.cantidadOut, "m2Out":dato.m2Out, "m2Conv":dato.m2Conv,"m2ConvWaste":dato.m2ConvWaste, "m2Corr":dato.m2Corr, "m2CorrPlanned":dato.m2CorrPlanned})

    for dato in Datos_MovPallets_B.objects.filter(programa=foto):
        labels2.append({"fechaini":dato.fechaini,"fechafin":dato.fechafin, "label": dato.label, "movscorr1":dato.movscorr1, "movscorr2":dato.movscorr2, "movsconv1":dato.movsconv1, "movsconv2":dato.movsconv2, "opcorr1":dato.opcorr1, "movscorrop1":dato.movscorrop1,"opcorr2":dato.opcorr2, "movscorrop2":dato.movscorrop2,"opcorr3":dato.opcorr3, "movscorrop3":dato.movscorrop3,"opcorr4":dato.opcorr4, "movscorrop4":dato.movscorrop4,"opcorr5":dato.opcorr5, "movscorrop5":dato.movscorrop5,"opcorr6":dato.opcorr6, "movscorrop6":dato.movscorrop6,"opcorr7":dato.opcorr7, "movscorrop7":dato.movscorrop7, "opconv1":dato.opconv1, "movsconvop1":dato.movsconvop1,"opconv2":dato.opconv2, "movsconvop2":dato.movsconvop2,"opconv3":dato.opconv3, "movsconvop3":dato.movsconvop3,"opconv4":dato.opconv4, "movsconvop4":dato.movsconvop4,"opconv5":dato.opconv5, "movsconvop5":dato.movsconvop5,"opconv6":dato.opconv6, "movsconvop6":dato.movsconvop6,"opconv7":dato.opconv7, "movsconvop7":dato.movsconvop7,"opconv8":dato.opconv8, "movsconvop8":dato.movsconvop8,"opconv9":dato.opconv9, "movsconvop9":dato.movsconvop9,"opconv10":dato.opconv10, "movsconvop10":dato.movsconvop10,"opconv11":dato.opconv11, "movsconvop11":dato.movsconvop11,"opconv12":dato.opconv12, "movsconvop12":dato.movsconvop12,"opconv13":dato.opconv13, "movsconvop13":dato.movsconvop13,"opconv14":dato.opconv14, "movsconvop14":dato.movsconvop14})

    data = {
    "labels":labels,
    "labels2":labels2,
            }
    print("Enviando datos movpallets")
    return JsonResponse(data)#http response con el datatype de JS



def get_data_inventario_prog(request, *args, **kwargs):
        #consulto en la base de datos todos los objetos pallet que tiene ubicación zTCY1 (a minúsculas). sumo sus m2 por pallets. los Cuento

            #veo de todos los order corrplan, cuál es la suma de m2 asociados a una máquina.



        print("iniciando consulta inv en Corrplan")


        invtotwip=0



        #for order in OrdenCorrplan.objects.filter(fecha_inicio__gte=ahora, fecha_inicio__lte=ahora+timedelta(days=8)):
        for pallet in Pallet.objects.filter( Q(ubic="ZFFG1") | Q(ubic="ZFFG2")| Q(ubic="ZFFW1")| Q(ubic="ZFFW2")| Q(ubic="ZDRO1")| Q(ubic="ZDRO2")| Q(ubic="ZTCY1") | Q(ubic="ZTCY2")| Q(ubic="ZWRD1")| Q(ubic="ZWRD2")| Q(ubic="ZHCR1")| Q(ubic="ZHCR2")| Q(ubic="ZSOB1") | Q(ubic="ZSOB2") | Q(ubic="ZPASILLO")):
        #        if pallet.ORDERID != order.order_id:
            invtotwip=invtotwip+pallet.m2pallet





        datosCorrplanINV={"TCY":{"m2Prog24h":DatosWIP_Prog.objects.get(maquina="TCY").m2Prog24h,"m2inv24h":DatosWIP_Prog.objects.get(maquina="TCY").m2inv24h,"m2inv":DatosWIP_Prog.objects.get(maquina="TCY").m2inv},"HCR":{"m2Prog24h":DatosWIP_Prog.objects.get(maquina="HCR").m2Prog24h,"m2inv24h":DatosWIP_Prog.objects.get(maquina="HCR").m2inv24h,"m2inv":DatosWIP_Prog.objects.get(maquina="HCR").m2inv},"WRD":{"m2Prog24h":DatosWIP_Prog.objects.get(maquina="WRD").m2Prog24h,"m2inv24h":DatosWIP_Prog.objects.get(maquina="WRD").m2inv24h,"m2inv":DatosWIP_Prog.objects.get(maquina="WRD").m2inv},"FFW":{"m2Prog24h":DatosWIP_Prog.objects.get(maquina="FFW").m2Prog24h,"m2inv24h":DatosWIP_Prog.objects.get(maquina="FFW").m2inv24h,"m2inv":DatosWIP_Prog.objects.get(maquina="FFW").m2inv},"DRO":{"m2Prog24h":DatosWIP_Prog.objects.get(maquina="DRO").m2Prog24h,"m2inv24h":DatosWIP_Prog.objects.get(maquina="DRO").m2inv24h,"m2inv":DatosWIP_Prog.objects.get(maquina="DRO").m2inv},"FFG":{"m2Prog24h":DatosWIP_Prog.objects.get(maquina="FFG").m2Prog24h,"m2inv24h":DatosWIP_Prog.objects.get(maquina="FFG").m2inv24h,"m2inv":DatosWIP_Prog.objects.get(maquina="FFG").m2inv}}

        print(datosCorrplanINV)

        fotocorrplan = FotoCorrplan.objects.all()[0].fecha_foto.strftime("%m/%d/%Y %H:%M:%S")
        data = {
        "datosCorrplanINV": datosCorrplanINV,
        "fotocorrplan": fotocorrplan,
        "invtotwip": invtotwip,

        #"filtroentrada": filtroentrada,
        #"filtrosalida": filtrosalida,
        #"filtroprod": filtroprod,

        }
        print("Enviando datos inventario")
        return JsonResponse(data)#http response con el datatype de JS



def get_data_inventario(request, *args, **kwargs):

     #consulto en la base de datos todos los objetos pallet que tiene ubicación zTCY1 (a minúsculas). sumo sus m2 por pallets. los Cuento

        prueba={"prueba1":(33,323), "prueba2":{"A":3,"B":4}}

        m2totalINV=0
        npalletstotalINV=0

        m2totalCORR=0
        npalletstotalCORR=0

        '''
        m2maqruta={"TCY":0, "HCR":0, "WRD":0, "FFG":0, "DRO":0, "FFW":0, "Otros":0}

        for ruta in m2maqruta.keys():

            filt= Pallet.objects.filter(maqruta=str(ruta)).filter(Q(ubic="ZPNC") | Q(ubic="ZHCR1") | Q(ubic="ZHCR2")| Q(ubic="ZTCY1")| Q(ubic="ZTCY2")| Q(ubic="ZWRD1")| Q(ubic="ZWRD2")| Q(ubic="ZSOB1")| Q(ubic="ZSOB2")| Q(ubic="ZFFW1")| Q(ubic="ZFFW2")| Q(ubic="ZDRO1")| Q(ubic="ZDRO2")| Q(ubic="ZFFG1")| Q(ubic="ZFFG2")| Q(ubic="ZPNC")| Q(ubic="ZPASILLO") )
            totm2=0
            for pallet in filt:
                totm2=totm2 + pallet.m2pallet
            m2maqruta[ruta]= totm2


        '''

        ##################################



        filtroentrada=[]
        filtrosalida=[]

        #Selecciono la foto más reciente:
        fotoinv=Foto_Datos_Inv_WIP.objects.all().order_by('-pk')[0]


        m2maqruta={}
        #Lo nuevo:
        for maqruta in m2Maqruta_WIP.objects.filter(programa=fotoinv):
            m2maqruta[maqruta.maquina]=maqruta.m2


        for mov in FiltroEntradaWIP.objects.filter(programa=fotoinv):
            #movimiento=[tarja, destino, hora]
            movimiento=[mov.LOADID, mov.ORDERID, mov.SOURCE, mov.DESTINATION, mov.EVENTDATETIME]
            filtroentrada.append(movimiento)

        for mov in FiltroSalidaWIP.objects.filter(programa=fotoinv):
            #movimiento=[tarja, destino, hora]
            movimiento=[mov.LOADID, mov.ORDERID, mov.SOURCE, mov.DESTINATION, mov.EVENTDATETIME]
            filtrosalida.append(movimiento)

        datosWIP={}
        for dato in Datos_Inv_WIP.objects.filter(programa=fotoinv):

            datosWIP.update( {dato.sector:{"cuenta":dato.cuenta,"m2tot":dato.m2tot,"indice":dato.indice,"dias":dato.dias,"al1":dato.al1,"al2":dato.al2,"al3":dato.al3} })


        data = {
        "m2maqruta":m2maqruta,
        "datosWIP":datosWIP,
        "m2totalINV": fotoinv.m2totalINV,
        "npalletstotalINV": npalletstotalINV,
        "m2totalCORR": m2totalCORR,
        "filtroentrada": filtroentrada,
        "filtrosalida": filtrosalida,
        #"filtroprod": filtroprod,

        }
        print("Enviando datos inventario")
        return JsonResponse(data)#http response con el datatype de JS


@csrf_exempt
def get_data_inv_ciclico(request, *args, **kwargs):


        #consulto en la base de datos todos los objetos pallet que tiene ubicación zTCY1 (a minúsculas). sumo sus m2 por pallets. los Cuento


        #prueba={"prueba1":(33,323), "prueba2":{"A":3,"B":4}}

        tomainv=TomaInvCic.objects.all().order_by('-pk')[0]
        ultimatoma=(tomainv.fechatomainvcic).strftime("%d/%m/%Y %H:%M:%S")


        '''
        datosWIP={"ZFFG1":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZFFG2":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZDRO1":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZDRO2":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZFFW1":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZFFW2":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZSOB1":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZWRD1":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZWRD2":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZSOB2":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZHCR1":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZHCR2":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZTCY1":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZTCY2":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZPNC":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0},"ZPASILLO":{"cuentaenc":0,"cuentanoenc":0,"cuentacti":0}}





        for calle in datosWIP.keys():

            palletstomainv = PalletCic.objects.filter(ubic=calle, tomainvcic=tomainv)

            #los pallets que están en esa ubicación según CTI, pero exluyendo los que ya están en palletstomainv
            palletscti= Pallet.objects.filter(ubic=calle).exclude(tarja__in=[o.tarja for o in palletstomainv]).order_by('tarja')

            #palletsnoencontrados son los que se pistolearon pero no aparecen en palletsCTI
            palletsnoencontrados=  PalletCic.objects.filter(ubic=calle, tomainvcic=tomainv).exclude(tarja__in=[o.tarja for o in Pallet.objects.filter(ubic=calle)]).order_by('tarja')

            #hago grupo de calles a excluir
            palletsCTIenotracalle = Pallet.objects.all().exclude(ubic=calle).order_by('tarja')
            tarjasnot=[]

            for aux in palletsCTIenotracalle:
                tarjasnot.append(aux.tarja)

            palletsenotracalle=[[],[]]

            for aux in PalletCic.objects.filter(ubic=calle, tomainvcic=tomainv).order_by('tarja'):
                if aux.tarja in tarjasnot:
                    palletsenotracalle[0].append(aux.tarja)

                    try:
                        palletsenotracalle[1].append(Pallet.objects.filter(tarja=aux.tarja)[0].ORDERID)
                    except:
                        #print("hola")
                        palletsenotracalle[1].append("vacio")

            datosWIP[calle]['palletsenotracalle'] = palletsenotracalle


            palletsencontrados = PalletCic.objects.filter(ubic=calle, tomainvcic=tomainv).exclude(tarja__in=[o.tarja for o in palletsnoencontrados]).order_by('tarja')



            #palletsnoencontrados= palletsnoencontrados.exclude(tarja__in=[o.tarja for o in palletsenotracalle]).order_by('tarja')

            lista=[[],[]]

            for pallet in palletscti.filter(ubic=calle).order_by('tarja'):

                lista[0].append(pallet.tarja)

                try:
                    lista[1].append(Pallet.objects.filter(tarja=pallet.tarja)[0].ORDERID)
                except:
                    #print("hola")
                    lista[1].append("vacio")

            datosWIP[calle]['palletscti'] = lista

            lista=[[],[]]

            for pallet in palletsencontrados.filter(ubic=calle).order_by('tarja'):

                lista[0].append(pallet.tarja)

                try:
                    lista[1].append(Pallet.objects.filter(tarja=pallet.tarja)[0].ORDERID)
                except:
                    #print("hola")
                    lista[1].append("vacio")

            datosWIP[calle]['palletsencontrados'] = lista


            lista=[[],[]]

            for pallet in palletsnoencontrados:

                if (pallet.tarja not in palletsenotracalle[0]):
                    lista[0].append(pallet.tarja)

                    try:
                        lista[1].append(Pallet.objects.filter(tarja=pallet.tarja)[0].ORDERID)
                    except:
                        #print("hola")
                        lista[1].append("vacio")

            datosWIP[calle]['palletsnoencontrados'] = lista



        '''
        #### Lo nuevo
        #'''
        fotoinv=Foto_Inv_Cic_WIP.objects.all().order_by('pk')[0]

        datosWIP2={}


        for calle in Foto_Calles_Inv_Cic_WIP.objects.filter(foto=fotoinv):

            lista2=[[],[]]
            datosWIP2[str(calle)]={}
            datosWIP2[str(calle)]['palletscti']=[[],[]]
            datosWIP2[str(calle)]['palletsencontrados']=[[],[]]
            datosWIP2[str(calle)]['palletsnoencontrados']=[[],[]]
            datosWIP2[str(calle)]['palletsenotracalle']=[[],[]]

            for pallet in Foto_Palletscti_Inv_Cic_WIP.objects.filter(calle=calle):
                if pallet:
                    datosWIP2[str(calle)]['palletscti'][0].append(pallet.pallet)
                    datosWIP2[str(calle)]['palletscti'][1].append(pallet.ORDERID)
                    #print(pallet.pallet)
                    #print(pallet.ORDERID)

            for pallet in Foto_Palletsencontrados_Inv_Cic_WIP.objects.filter(calle=calle):
                if pallet:
                    datosWIP2[str(calle)]['palletsencontrados'][0].append(pallet.pallet)
                    datosWIP2[str(calle)]['palletsencontrados'][1].append(pallet.ORDERID)
                    #print(pallet.pallet)
                    #print(pallet.ORDERID)

            for pallet in Foto_Palletsenotracalle_Inv_Cic_WIP.objects.filter(calle=calle):
                if pallet:
                    datosWIP2[str(calle)]['palletsenotracalle'][0].append(pallet.pallet)
                    datosWIP2[str(calle)]['palletsenotracalle'][1].append(pallet.ORDERID)
                    #print(pallet.pallet)
                    #print(pallet.ORDERID)

            for pallet in Foto_Palletsnoencontrados_Inv_Cic_WIP.objects.filter(calle=calle):
                if pallet:
                    datosWIP2[str(calle)]['palletsnoencontrados'][0].append(pallet.pallet)
                    datosWIP2[str(calle)]['palletsnoencontrados'][1].append(pallet.ORDERID)
                    #print(pallet.pallet)
                    #print(pallet.ORDERID)


            #'''

        data = {
        "datosWIP":datosWIP2,
        "ultimatoma":ultimatoma,


        }
        print("Enviando datos inventario")
        return JsonResponse(data)#http response con el datatype de JS



def carga_mov_pallets(request):
    template_name = 'blog/carga_mov_pallets.html'

    if request.method == "POST":
        form = MovPalletForm(request.POST)

        if form.is_valid():

            dato0=form.cleaned_data["TRANSACTIONINDEX"]
            dato1=form.cleaned_data["PLANTID"]
            dato2=form.cleaned_data["WAREHOUSE"]
            dato3=form.cleaned_data["INTERNALSPECID"]
            dato4=form.cleaned_data["ORDERID"]
            dato5=form.cleaned_data["PARTID"]
            dato6=form.cleaned_data["OPERATIONNO"]
            dato7=form.cleaned_data["UNITTYPE"]
            dato8=form.cleaned_data["LOADID"]
            dato9=form.cleaned_data["UNITNO"]
            dato10=form.cleaned_data["SOURCE"]
            dato11=form.cleaned_data["DESTINATION"]
            dato12=form.cleaned_data["EVENTDATETIME"]
            #Acá proceso el dato12 para pasarlo a datetime y poder guardarlo en el modelo.
            dato13=form.cleaned_data["EVENTTIME"]
            datounidadespallet=form.cleaned_data["unidadespallet"]
            datokgpallet=form.cleaned_data["kgpallet"]
            datom2pallet=form.cleaned_data["m2pallet"]
            datoalto=form.cleaned_data["alto"]
            datoancho=form.cleaned_data["ancho"]
            datokguni=form.cleaned_data["kguni"]
            datom2uni=form.cleaned_data["m2uni"]
            datoFGLoad=form.cleaned_data["esFGLoad"]

            #Ojo aquí si cambia algún dato en un transactionindex lo va a duplicar?
            o, created = MovPallets.objects.get_or_create(TRANSACTIONINDEX=dato0)
            o.PLANTID=dato1
            o.WAREHOUSE=dato2
            o.INTERNALSPECID=dato3
            o.ORDERID=dato4
            o.PARTID=dato5
            o.OPERATIONNO=dato6
            o.UNITTYPE=dato7
            o.LOADID=dato8
            o.UNITNO=dato9
            o.SOURCE=dato10
            o.DESTINATION=dato11
            o.EVENTDATETIME=dato12
            o.EVENTTIME=dato13
            o.unidadespallet=datounidadespallet
            o.kgpallet=datokgpallet
            o.m2pallet=datom2pallet
            o.alto=datoalto
            o.ancho=datoancho
            o.kguni=datokguni
            o.m2uni=datom2uni
            o.esFGLoad=datoFGLoad

            o.save()

            #creo la ubicación de inventario en caso de que no exista.
            a, created = UbicPallet.objects.get_or_create(calle=dato11)
            a.save()

            #creo el pallet en caso de que no exista. Si ya existe le actualizo la ubicación.

            c, created = Pallet.objects.get_or_create(tarja=dato8)
            c.padron=dato3
            c.ubic=dato11
            c.ORDERID=dato4
            #c.ubic2=UbicPallet.objects.filter(calle=dato11)[0]
            if datoFGLoad==1: #Aquí hay que arreglar el caso en que se mueve un pallet que todavía no está creado*ya está considerado, usa la fecha now() por defecto
                c.fechacreac=dato12
            c.ancho=datoancho
            c.alto=datoalto
            c.unidades=datounidadespallet
            c.cliente="pendiente"
            c.m2uni=datom2uni
            c.kguni=datokguni
            c.m2pallet=datom2pallet
            c.kgpallet=datokgpallet
            c.save()

        else:

            dato0=form.cleaned_data["TRANSACTIONINDEX"]
            dato1=form.cleaned_data["PLANTID"]
            dato2=form.cleaned_data["WAREHOUSE"]
            dato3=form.cleaned_data["INTERNALSPECID"]
            dato4=form.cleaned_data["ORDERID"]
            dato5=form.cleaned_data["PARTID"]
            dato6=form.cleaned_data["OPERATIONNO"]
            dato7=form.cleaned_data["UNITTYPE"]
            dato8=form.cleaned_data["LOADID"]
            dato9=form.cleaned_data["UNITNO"]
            dato10=form.cleaned_data["SOURCE"]
            dato11=form.cleaned_data["DESTINATION"]
            dato12=form.cleaned_data["EVENTDATETIME"]
            #Acá proceso el dato12 para pasarlo a datetime y poder guardarlo en el modelo.
            dato13=form.cleaned_data["EVENTTIME"]
            datounidadespallet=form.cleaned_data["unidadespallet"]
            datokgpallet=form.cleaned_data["kgpallet"]
            datom2pallet=form.cleaned_data["m2pallet"]
            datoalto=form.cleaned_data["alto"]
            datoancho=form.cleaned_data["ancho"]
            datokguni=form.cleaned_data["kguni"]
            datom2uni=form.cleaned_data["m2uni"]
            datoFGLoad=form.cleaned_data["esFGLoad"]

            #Ojo aquí si cambia algún dato en un transactionindex lo va a duplicar?
            o, created = MovPallets.objects.get_or_create(TRANSACTIONINDEX=dato0)
            o.PLANTID=dato1
            o.WAREHOUSE=dato2
            o.INTERNALSPECID=dato3
            o.ORDERID=dato4
            o.PARTID=dato5
            o.OPERATIONNO=dato6
            o.UNITTYPE=dato7
            o.LOADID=dato8
            o.UNITNO=dato9
            o.SOURCE=dato10
            o.DESTINATION=dato11
            o.EVENTDATETIME=dato12
            o.EVENTTIME=dato13
            o.unidadespallet=datounidadespallet
            o.kgpallet=datokgpallet
            o.m2pallet=datom2pallet
            o.alto=datoalto
            o.ancho=datoancho
            o.kguni=datokguni
            o.m2uni=datom2uni
            o.esFGLoad=datoFGLoad

            o.save()

            #creo la ubicación de inventario en caso de que no exista.
            a, created = UbicPallet.objects.get_or_create(calle=dato11)
            a.save()

            #creo el pallet en caso de que no exista. Si ya existe le actualizo la ubicación.

            c, created = Pallet.objects.get_or_create(tarja=dato8)
            c.padron=dato3
            c.ubic=dato11
            c.ORDERID=dato4
            #c.ubic2=UbicPallet.objects.filter(calle=dato11)[0]
            if datoFGLoad==1: #Aquí hay que arreglar el caso en que se mueve un pallet que todavía no está creado*ya está considerado, usa la fecha now() por defecto
                c.fechacreac=dato12
            c.ancho=datoancho
            c.alto=datoalto
            c.unidades=datounidadespallet
            c.cliente="pendiente"
            c.m2uni=datom2uni
            c.kguni=datokguni
            c.m2pallet=datom2pallet
            c.kgpallet=datokgpallet
            c.save()

            #form = MovPalletForm()#Esto se pone si quieres que después de submitear, los valores que pusiste en los form se borren

            #form = MovPalletForm()


        #return redirect ('res_inventario')

    else:
        form = MovPalletForm()

    #bobinas = reversed(list(BobInvCic.objects.all()))

    ultimotransaction=str(MovPallets.objects.all().order_by('-TRANSACTIONINDEX')[0].TRANSACTIONINDEX)

    #print(mensajes)

    return render(request, template_name, {'form':form, "ultimo":ultimotransaction})#,'bobinas':bobinas,"mensajes":mensajes})


def invsimple2(request):
    template_name = 'blog/invsimple2.html'

    datocrudo='I'
    tomainv=TomaInvCic.objects.all().order_by('-pk')[0]

    if request.method == "POST":
        form = PruebaModForm(request.POST)


        datocrudo=form.data["dato1"]
        initial = {}
        form = PruebaModForm(initial=initial)



        if datocrudo[:1].upper()=="Z":
            tomainv.aux1=datocrudo.upper()
            tomainv.save()
            '''

            '''
        elif len(datocrudo) <= 7 and len(datocrudo) > 5 and datocrudo.isdigit():
            o, created = PalletCic.objects.get_or_create(tarja=datocrudo, tomainvcic=tomainv)
            o.ubic= tomainv.aux1.upper()
            o.save()




        #return redirect ('res_inventario')

    else:
        #initial = {'name': 'Initial name'}

        initial = {}
        form = PruebaModForm(initial=initial)


    #bobinas = reversed(list(BobInvCic.objects.all()))



    palletstomainv = PalletCic.objects.filter(ubic=tomainv.aux1.upper(), tomainvcic=tomainv)
    #los pallets que están en esa ubicación según CTI, pero exluyendo los que ya están en palletstomainv
    palletscti= Pallet.objects.filter(ubic=tomainv.aux1.upper()).exclude(tarja__in=[o.tarja for o in palletstomainv]).order_by('pk')

    #palletsnoencontrados son los que se pistolearon pero no aparecen en palletsCTI
    palletsnoencontrados=  PalletCic.objects.filter(ubic=tomainv.aux1.upper(), tomainvcic=tomainv).exclude(tarja__in=[o.tarja for o in Pallet.objects.filter(ubic=tomainv.aux1.upper())]).order_by('pk')

    palletsencontrados = PalletCic.objects.filter(ubic=tomainv.aux1.upper(), tomainvcic=tomainv).exclude(tarja__in=[o.tarja for o in palletsnoencontrados]).order_by('pk')








    return render(request, template_name, {'form':form, 'palletsnoencontrados':palletsnoencontrados, 'palletsencontrados':palletsencontrados, 'tomainv':tomainv,'palletscti':palletscti, 'cuentaenc': palletsencontrados.count(), 'cuentanoenc': palletsnoencontrados.count(), 'cuentacti':palletscti.count()})



def invsimple(request):
    template_name = 'blog/invsimple.html'

    if request.method == "POST":
        form = PruebaModForm(request.POST)

        if form.is_valid():


            datocrudo=form.cleaned_data["dato1"]
            datocrudo2=form.cleaned_data["dato2"]
            o, created = obInvCic.objects.get_or_create(nbobina=datocrudo)
            o.ubic=datocrudo2
            o.save()

            print(datocrudo)
            form = PruebaModForm()


        else:
            datocrudo=form.data["dato1"]
            datocrudo2=form.cleaned_data["dato2"]
            o , created = BobInvCic.objects.get_or_create(nbobina=datocrudo)
            o.ubic=datocrudo2
            o.save()

            print(datocrudo)
            o=BobInvCic.objects.all().order_by('-pk')[0]

            initial = {'dato2': o.ubic}
            form = PruebaModForm(initial=initial)


        #return redirect ('res_inventario')

    else:
        #initial = {'name': 'Initial name'}
        o=BobInvCic.objects.all().order_by('-pk')[0]
        initial = {'dato2': o.ubic}
        form = PruebaModForm(initial=initial)


    bobinas = reversed(list(BobInvCic.objects.all()))

    mensajes=str(BobInvCic.objects.all().count())

    print(mensajes)

    return render(request, template_name, {'form':form,'bobinas':bobinas,"mensajes":mensajes})




def panel_bpt(request):
    #print("cargando consumos puestos")
    template_name = 'blog/panel_bpt.html'


    return render(request, template_name, {})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?

def panel_wip(request):
    #print("cargando consumos puestos")
    template_name = 'blog/panel_wip.html'


    return render(request, template_name, {})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?

def panel_wip_prog(request):
    #print("cargando consumos puestos")
    template_name = 'blog/panel_wip_prog.html'


    return render(request, template_name, {})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?





class UbicDetailView(DetailView):
    context_object_name = 'ubicpallet_detail'
    model = UbicPallet

    def actualizadatos(self, pk):
        #referencia = OrdenProg.objects.filter(pk=pk[0])#ojo tmabién puede ser con el get. ahí hay que poner el [0] ya que no acepta más de un resultado
        referencia = OrdenProg.objects.get(pk=pk)
        #detalles = DetalleProg.objects.filter(programma = referencia)
        #ordenprevia = OrdenProg.objects.get(pk=pk).get_previous_by_fecha_programa()



    def get_context_data(self, **kwargs):

        pk = self.kwargs['pk']# this is the primary key from your URL
        # your other code

        print(pk)

        #self.actualizadatos(pk)

        context = super().get_context_data(**kwargs)
        #detalleprog2 = DetalleProg.objects.filter(programma = OrdenProg.objects.get(pk=pk), datefinajustada__lte = OrdenProg.objects.get(pk=pk).horizontefin )#.filter(published_date__isnull=True).order_by('-published_date')

        #detalleprogprev = DetalleProg.objects.filter(programma = OrdenProg.objects.get(pk=pk).get_previous_by_fecha_programa(), datefin__lt= OrdenProg.objects.get(pk=pk).fecha_programa, datefinajustada__gte = OrdenProg.objects.get(pk=pk).fecha_programa.date() )

        #context['detalleprog'] = detalleprog2.union(detalleprogprev, all=True).order_by('datefin')

        #context['prodreal'] = ProdReal.objects.filter(datefin__gte=OrdenProg.objects.get(pk=pk).fecha_programa + timedelta(days=-1), datefin__lt=OrdenProg.objects.get(pk=pk).horizontefin + timedelta(days=1)).order_by('datefin')#Acá hay que filtrar para que sean las órdenes reales desde la fecha del programa de referencia en adelante
        #context['maquinas'] = Maquinas.objects.all()#.filter(published_date__isnull=True).order_by('-published_date')
        #context['turnos'] = Turnos.objects.all()#.filter(published_date__isnull=True).order_by('-published_date')
        #context['orderinfos'] = OrderInfo.objects.all() #filtrar para que sólo mande los que están dentro del detalleprog?
        #context['padrones'] = Padron.objects.all()#Agregarle al orderinfo fecha de subida y filtrar para que busque?
        context['ubic'] = UbicPallet.objects.get(pk=pk)
        context['m2tot'] = str( round(UbicPallet.objects.get(pk=pk).m2tot, 1) )
        context['npallets'] = str(UbicPallet.objects.get(pk=pk).npallets)
        context['pallets'] = Pallet.objects.filter(ubic__iexact=str(UbicPallet.objects.get(pk=pk).calle)).order_by("ORDERID")
        return context



def consumos_puestos(request):
    print("cargando consumos puestos")
    template_name = 'blog/consumos_puestos.html'


    return render(request, template_name, {})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?




def get_data_corrplan_cartones(request, *args, **kwargs):


    #datefinajustadaobj2 = datefinajustadaini.strftime("%d %m %Y")


    print("cargando corrplan")


            ## [N° máquina, n° piezas, Area]
        #Actualizo el objeto InfoWIP n° cero:
    #genero el diccionario con la suma de todas las àreas y los tipos de cartòn




    datos0=[324,42,23,234,4]

    #datos9=model_to_dict( InfoWIP.objects.all().order_by('contador')[9])#["red","blue","green"]//EL RESTO DE LOS DATOS LOS SACO DE LA BASE DE DATOS?


    labels=[]#agrego las fechas de cada dìa que quiero analizar (5 a partir de mañana)
    metros2FFG=[]
    metros2FFW=[]
    metros2TCY=[]
    metros2DRO=[]
    metros2WRD=[]
    metros2HCR=[]
    metros2tot=[]
    #append fecha hoy +1

    horizonte=10
    ahora=datetime.now()

    for i in range(0,horizonte):
        labels.append((ahora+timedelta(days=i)).replace(hour= 0, minute=0, second=0, microsecond=0))
        metros2FFG.append(0)
        metros2FFW.append(0)
        metros2TCY.append(0)
        metros2DRO.append(0)
        metros2WRD.append(0)
        metros2HCR.append(0)
        metros2tot.append(0)


    default_items=[123, 124, 432]

    fotocorr= FotoCorrplan.objects.all().order_by("-fecha_foto")[0]
    fecha_foto=fotocorr.fecha_foto
    print(fecha_foto)



    FFG= Maquinas.objects.filter(maquina="FFG")[0]
    FFW= Maquinas.objects.filter(maquina="FFW")[0]
    TCY= Maquinas.objects.filter(maquina="TCY")[0]
    DRO= Maquinas.objects.filter(maquina="DRO")[0]
    WRD= Maquinas.objects.filter(maquina="WRD")[0]
    HCR= Maquinas.objects.filter(maquina="HCR")[0]


    #grabo suma de m2 diarios que coinciden con las fechas fijadas en el arreglo labels2
    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, maquina= FFG, fecha_inicio__gt=((ahora).replace(hour= 0, minute=0, second=0, microsecond=0)), fecha_inicio__lte=((ahora+timedelta(days=horizonte)).replace(hour= 0, minute=0, second=0, microsecond=0)) )
    for orden in ordenes:

        for i in range(0,horizonte):
            if (orden.fecha_inicio).replace(hour= 0, minute=0, second=0, microsecond=0) == labels[i]:
                metros2FFG[i]=metros2FFG[i]+orden.area

    #grabo suma de m2 diarios que coinciden con las fechas fijadas en el arreglo labels2
    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, maquina= FFW, fecha_inicio__gt=((ahora).replace(hour= 0, minute=0, second=0, microsecond=0)), fecha_inicio__lte=((ahora+timedelta(days=horizonte)).replace(hour= 0, minute=0, second=0, microsecond=0)) )
    for orden in ordenes:

        for i in range(0,horizonte):
            if (orden.fecha_inicio).replace(hour= 0, minute=0, second=0, microsecond=0) == labels[i]:
                metros2FFW[i]=metros2FFW[i]+orden.area


    #grabo suma de m2 diarios que coinciden con las fechas fijadas en el arreglo labels2
    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, maquina= TCY, fecha_inicio__gt=((ahora).replace(hour= 0, minute=0, second=0, microsecond=0)), fecha_inicio__lte=((ahora+timedelta(days=horizonte)).replace(hour= 0, minute=0, second=0, microsecond=0)) )
    for orden in ordenes:

        for i in range(0,horizonte):
            if (orden.fecha_inicio).replace(hour= 0, minute=0, second=0, microsecond=0) == labels[i]:
                metros2TCY[i]=metros2TCY[i]+orden.area


    #grabo suma de m2 diarios que coinciden con las fechas fijadas en el arreglo labels2
    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, maquina= DRO, fecha_inicio__gt=((ahora).replace(hour= 0, minute=0, second=0, microsecond=0)), fecha_inicio__lte=((ahora+timedelta(days=horizonte)).replace(hour= 0, minute=0, second=0, microsecond=0)) )
    for orden in ordenes:

        for i in range(0,horizonte):
            if (orden.fecha_inicio).replace(hour= 0, minute=0, second=0, microsecond=0) == labels[i]:
                metros2DRO[i]=metros2DRO[i]+orden.area


    #grabo suma de m2 diarios que coinciden con las fechas fijadas en el arreglo labels2
    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, maquina= WRD, fecha_inicio__gt=((ahora).replace(hour= 0, minute=0, second=0, microsecond=0)), fecha_inicio__lte=((ahora+timedelta(days=horizonte)).replace(hour= 0, minute=0, second=0, microsecond=0)) )
    for orden in ordenes:

        for i in range(0,horizonte):
            if (orden.fecha_inicio).replace(hour= 0, minute=0, second=0, microsecond=0) == labels[i]:
                metros2WRD[i]=metros2WRD[i]+orden.area

    #grabo suma de m2 diarios que coinciden con las fechas fijadas en el arreglo labels2
    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, maquina= HCR, fecha_inicio__gt=((ahora).replace(hour= 0, minute=0, second=0, microsecond=0)), fecha_inicio__lte=((ahora+timedelta(days=horizonte)).replace(hour= 0, minute=0, second=0, microsecond=0)) )
    for orden in ordenes:

        for i in range(0,horizonte):
            if (orden.fecha_inicio).replace(hour= 0, minute=0, second=0, microsecond=0) == labels[i]:
                metros2HCR[i]=metros2HCR[i]+orden.area




    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr,fecha_inicio__gt=((ahora).replace(hour= 0, minute=0, second=0, microsecond=0)), fecha_inicio__lte=((ahora+timedelta(days=horizonte)).replace(hour= 0, minute=0, second=0, microsecond=0)) )
    for orden in ordenes:

        for i in range(0,horizonte):
            if (orden.fecha_inicio).replace(hour= 0, minute=0, second=0, microsecond=0) == labels[i]:
                metros2tot[i]=metros2tot[i]+orden.area

    print("metros2tot: ")
    print(metros2tot[0])
    print(metros2FFG[0])
    print(metros2FFW[0])
    print(metros2TCY[0])
    print(metros2DRO[0])
    print(metros2WRD[0])
    print(metros2HCR[0])


    data = {
    "labels": labels,
    "defaults": default_items,
    "datos0": datos0,
    "metros2FFG": metros2FFG,
    "metros2FFW": metros2FFW,
    "metros2TCY": metros2TCY,
    "metros2DRO": metros2DRO,
    "metros2WRD": metros2WRD,
    "metros2HCR": metros2HCR,

    "fecha_foto": fecha_foto,

    }
    print("Enviando Json Datos Graph")
    return JsonResponse(data)#http response con el datatype de JS







def get_corrplan(request):
    #print("cargando datos wip")
    template_name = 'blog/get_corrplan.html'

    corrplan= [["hola"], ["q ase"]]
    corrplan = webscrap2.webscrap_corrplan()[2]

    print(webscrap2.webscrap_corrplan()[2])
    return render(request, template_name, {'corrplan':corrplan,})#acá le puedo decir que los mande ordenados por fecha?




class CamionDetailView(DetailView):
    context_object_name = 'camion_detail'
    model = Camion

    def actualizadatos(self, pk):
        #referencia = OrdenProg.objects.filter(pk=pk[0])#ojo tmabién puede ser con el get. ahí hay que poner el [0] ya que no acepta más de un resultado
        referencia = OrdenProg.objects.get(pk=pk)
        #detalles = DetalleProg.objects.filter(programma = referencia)
        #ordenprevia = OrdenProg.objects.get(pk=pk).get_previous_by_fecha_programa()



    def get_context_data(self, **kwargs):

        pk = self.kwargs['pk']# this is the primary key from your URL
        # your other code

        print(pk)

        #self.actualizadatos(pk)

        context = super().get_context_data(**kwargs)
        #detalleprog2 = DetalleProg.objects.filter(programma = OrdenProg.objects.get(pk=pk), datefinajustada__lte = OrdenProg.objects.get(pk=pk).horizontefin )#.filter(published_date__isnull=True).order_by('-published_date')

        #detalleprogprev = DetalleProg.objects.filter(programma = OrdenProg.objects.get(pk=pk).get_previous_by_fecha_programa(), datefin__lt= OrdenProg.objects.get(pk=pk).fecha_programa, datefinajustada__gte = OrdenProg.objects.get(pk=pk).fecha_programa.date() )

        #context['detalleprog'] = detalleprog2.union(detalleprogprev, all=True).order_by('datefin')

        #context['prodreal'] = ProdReal.objects.filter(datefin__gte=OrdenProg.objects.get(pk=pk).fecha_programa + timedelta(days=-1), datefin__lt=OrdenProg.objects.get(pk=pk).horizontefin + timedelta(days=1)).order_by('datefin')#Acá hay que filtrar para que sean las órdenes reales desde la fecha del programa de referencia en adelante
        #context['maquinas'] = Maquinas.objects.all()#.filter(published_date__isnull=True).order_by('-published_date')
        #context['turnos'] = Turnos.objects.all()#.filter(published_date__isnull=True).order_by('-published_date')
        #context['orderinfos'] = OrderInfo.objects.all() #filtrar para que sólo mande los que están dentro del detalleprog?
        #context['padrones'] = Padron.objects.all()#Agregarle al orderinfo fecha de subida y filtrar para que busque?
        context['camiones'] = Camion.objects.get(pk=pk)
        context['patenteqr'] = Camion.objects.get(pk=pk).Patente
        context['infoqr'] = (str(Camion.objects.get(pk=pk).Chofer) + "\t" + str(Camion.objects.get(pk=pk).Telefono) + "\t" + "\t" + str(Camion.objects.get(pk=pk).Rut) + "\t" + str(Camion.objects.get(pk=pk).Transportista) )
        return context



def qr_despacho(request):
    #print("cargando datos wip")
    template_name = 'blog/qr_despacho.html'

    camiones = Camion.objects.all() #Post.objects.filter(published_date__isnull=True).order_by('create_date')
    transportistas = []
    for camion in camiones:
        if not camion.Transportista in transportistas:
            transportistas.append(camion.Transportista)




    return render(request, template_name, {'camiones':camiones, 'transportistas':transportistas})#acá le puedo decir que los mande ordenados por fecha?

#para crear new camiòn

def new_camion(request):

    if request.method == "POST":
        form = QRForm(request.POST)
        if form.is_valid():
            camion = form.save(commit=False)


            #post.published_date = timezone.now()
            camion.save()
            return redirect('camion_detail', pk=camion.pk)
    else:

        form = QRForm()

    return render(request, 'blog/camion_edit.html', {'form': form})


def placas_wip(request):
    print("cargando datos wip")
    template_name = 'blog/placas_wip.html'


    return render(request, template_name, {})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?



def get_data_wip(request, *args, **kwargs):#este hace que se actualice utilizando el wescrap, el otro no (mejor que no y se actualiza de manera coordinada a fin de cada turno.)

    #datefinajustadaobj2 = datefinajustadaini.strftime("%d %m %Y")


    print("cargando datos wip")

    placas_wip = pruebawebscrap.webscrap_wip()
            ## [N° máquina, n° piezas, Area]
    print("preparandose para copiar")
    for i in range(9,0,-1):
        print("copiando datos "+ str(i))

        Info=InfoWIP.objects.all().order_by('contador')[i]
        Infoprev=InfoWIP.objects.all().order_by('contador')[i-1]


        Info.M2FFG=Infoprev.M2FFG
        Info.PiFFG=Infoprev.PiFFG

        Info.M2FFW=Infoprev.M2FFW
        Info.PiFFW=Infoprev.PiFFW

        Info.M2TCY=Infoprev.M2TCY
        Info.PiTCY=Infoprev.PiTCY

        Info.M2DRO=Infoprev.M2DRO
        Info.PiDRO=Infoprev.PiDRO

        Info.M2WRD=Infoprev.M2WRD
        Info.PiWRD=Infoprev.PiWRD

        Info.M2HCR=Infoprev.M2HCR
        Info.PiHCR=Infoprev.PiHCR

        Info.M2DIM=Infoprev.M2DIM
        Info.PiDIM=Infoprev.PiDIM

        Info.M2CORR=Infoprev.M2CORR
        Info.PiCORR=Infoprev.PiCORR

        Info.M2Total= Infoprev.M2Total
        Info.PiTotal= Infoprev.PiTotal

        Info.save()

    #Actualizo el objeto InfoWIP n° cero:
    Info=InfoWIP.objects.all().order_by('contador')[0]

    Info.M2FFG=round(placas_wip[0][2],2)
    Info.PiFFG=round(placas_wip[0][1],2)

    Info.M2FFW=round(placas_wip[2][2],2)
    Info.PiFFW=round(placas_wip[2][1],2)

    Info.M2TCY=round(placas_wip[1][2],2)
    Info.PiTCY=round(placas_wip[1][1],2)

    Info.M2DRO=round(placas_wip[3][2],2)
    Info.PiDRO=round(placas_wip[3][1],2)

    Info.M2WRD=round(placas_wip[4][2],2)
    Info.PiWRD=round(placas_wip[4][1],2)

    Info.M2HCR=round(placas_wip[5][2],2)
    Info.PiHCR=round(placas_wip[5][1],2)

    Info.M2DIM=round(placas_wip[6][2],2)
    Info.PiDIM=round(placas_wip[6][1],2)

    Info.M2CORR=round(placas_wip[7][2],2)
    Info.PiCORR=round(placas_wip[7][1],2)

    Info.M2Total= Info.M2FFG +Info.M2FFW+Info.M2TCY+Info.M2DRO+Info.M2WRD+Info.M2HCR+Info.M2DIM+Info.M2CORR
    Info.PiTotal= Info.PiFFG +Info.PiFFW+Info.PiTCY+Info.PiDRO+Info.PiWRD+Info.PiHCR+Info.PiDIM+Info.PiCORR


    Info.save()




    auxsumapiezas=0
    auxsumaarea=0

    for i in range(0,len(placas_wip)):
        auxsumapiezas=auxsumapiezas+placas_wip[i][1]
        auxsumaarea=auxsumaarea+placas_wip[i][2]

    totales=[auxsumapiezas,auxsumaarea]
    print("carga datos completada")


    #datos0=placas_wip#["red","blue","green"]
    datos0=model_to_dict( InfoWIP.objects.all().order_by('contador')[0])
    datos1=model_to_dict( InfoWIP.objects.all().order_by('contador')[1])
    datos2=model_to_dict( InfoWIP.objects.all().order_by('contador')[2])
    datos3=model_to_dict( InfoWIP.objects.all().order_by('contador')[3])
    datos4=model_to_dict( InfoWIP.objects.all().order_by('contador')[4])
    datos5=model_to_dict( InfoWIP.objects.all().order_by('contador')[5])
    datos6=model_to_dict( InfoWIP.objects.all().order_by('contador')[6])
    datos7=model_to_dict( InfoWIP.objects.all().order_by('contador')[7])
    datos8=model_to_dict( InfoWIP.objects.all().order_by('contador')[8])
    datos9=model_to_dict( InfoWIP.objects.all().order_by('contador')[9])#["red","blue","green"]//EL RESTO DE LOS DATOS LOS SACO DE LA BASE DE DATOS?
    totales=totales


    ahora=datetime.now()
    m2pormaq=[] #agrego el dato de los m2 programados por màquina que hay en las pròximas 24h.
    fotocorr= FotoCorrplan.objects.all().order_by("-fecha_foto")[0]
    #print(fotocorr.fecha_foto)

    FFG= Maquinas.objects.filter(maquina="FFG")
    FFW= Maquinas.objects.filter(maquina="FFW")
    TCY= Maquinas.objects.filter(maquina="TCY")
    DRO= Maquinas.objects.filter(maquina="DRO")
    WRD= Maquinas.objects.filter(maquina="WRD")
    HCR= Maquinas.objects.filter(maquina="HCR")
    #grabo suma de m2 diarios que coinciden con las fechas fijadas en el arreglo labels2



    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=FFG ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)

    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=FFW ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)

    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=TCY ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)

    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=DRO ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)

    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=WRD ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)

    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=HCR ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)



    data = {
    "datos0": datos0,
    "datos1": datos1,
    "datos2": datos2,
    "datos3": datos3,
    "datos4": datos4,
    "datos5": datos5,
    "datos6": datos6,
    "datos7": datos7,
    "datos8": datos8,
    "datos9": datos9,
    "totales": totales,
    "m2pormaq": m2pormaq,
    }
    print("Enviando Json Datos Graph")
    return JsonResponse(data)#http response con el datatype de JS




def get_data_wip_2(request, *args, **kwargs):


    #datefinajustadaobj2 = datefinajustadaini.strftime("%d %m %Y")


    print("cargando datos wip")


            ## [N° máquina, n° piezas, Area]
        #Actualizo el objeto InfoWIP n° cero:




    #datos0=placas_wip#["red","blue","green"]
    datos0=model_to_dict( InfoWIP.objects.all().order_by('contador')[0])
    datos1=model_to_dict( InfoWIP.objects.all().order_by('contador')[1])
    datos2=model_to_dict( InfoWIP.objects.all().order_by('contador')[2])
    datos3=model_to_dict( InfoWIP.objects.all().order_by('contador')[3])
    datos4=model_to_dict( InfoWIP.objects.all().order_by('contador')[4])
    datos5=model_to_dict( InfoWIP.objects.all().order_by('contador')[5])
    datos6=model_to_dict( InfoWIP.objects.all().order_by('contador')[6])
    datos7=model_to_dict( InfoWIP.objects.all().order_by('contador')[7])
    datos8=model_to_dict( InfoWIP.objects.all().order_by('contador')[8])
    datos9=model_to_dict( InfoWIP.objects.all().order_by('contador')[9])#["red","blue","green"]//EL RESTO DE LOS DATOS LOS SACO DE LA BASE DE DATOS?


    print("cargando m2 cargados por màquina")

    ahora=datetime.now()
    m2pormaq=[] #agrego el dato de los m2 programados por màquina que hay en las pròximas 24h.
    print("a1")
    fotocorr= FotoCorrplan.objects.all().order_by("-fecha_foto")[0]
    #print(fotocorr.fecha_foto)
    print("a2")
    FFG= Maquinas.objects.filter(maquina="FFG")[0]
    print("a3")
    FFW= Maquinas.objects.filter(maquina="FFW")[0]
    TCY= Maquinas.objects.filter(maquina="TCY")[0]
    DRO= Maquinas.objects.filter(maquina="DRO")[0]
    WRD= Maquinas.objects.filter(maquina="WRD")[0]
    HCR= Maquinas.objects.filter(maquina="HCR")[0]
    #grabo suma de m2 diarios que coinciden con las fechas fijadas en el arreglo labels2


    print("a4")
    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=FFG ).order_by('fecha_inicio')
    print("a5")
    sumam2=0
    for orden in ordenes:

        sumam2 = sumam2 + orden.area
    print(sumam2)
    m2pormaq.append(sumam2)

    print("a6")

    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=FFW ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)

    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=TCY ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)

    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=DRO ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)

    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=WRD ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)

    ordenes = OrdenCorrplan.objects.filter(programa= fotocorr, fecha_inicio__gt=((ahora)), fecha_inicio__lte=((ahora+timedelta(days=1))), maquina=HCR ).order_by('fecha_inicio')
    sumam2=0
    for orden in ordenes:
        sumam2 = sumam2 + orden.area
    m2pormaq.append(sumam2)


    print("a7")

    data = {
    "datos0": datos0,
    "datos1": datos1,
    "datos2": datos2,
    "datos3": datos3,
    "datos4": datos4,
    "datos5": datos5,
    "datos6": datos6,
    "datos7": datos7,
    "datos8": datos8,
    "datos9": datos9,
    "m2pormaq": m2pormaq,


    }
    print("Enviando Json Datos Graph")
    return JsonResponse(data)#http response con el datatype de JS





def cump_proto(request):

    print("carga datos iniciada")
    template_name = 'blog/cump_proto.html'

    ini= datetime.now()

    producciones = ProdRealCorr.objects.filter(datefinajustada__gte=ini.replace(hour= 0, minute=0, second=0, microsecond=0)-timedelta(days=50)).order_by('datefin') #Post.objects.filter(published_date__isnull=True).order_by('create_date')
    #turnos = Turnos.objects.all()
    maquinas = Maquinas.objects.all()
    # Trato de procesar toda la info en el server y solo mandar los datos que hay que poner en la tabla.
    dias = []
    MLprodreal = []
    trims = []
    for i in range(0,50):
        fecha=(datetime.now()-timedelta(days=i)).replace(hour= 0, minute=0, second=0, microsecond=0)
        dias.append( fecha.strftime("%d-%m-%y"))
        Prods=ProdRealCorr.objects.filter(datefinajustada=fecha).order_by('datefin')
        metros=0
        trimpond=0

        print ("Ok listo para imprimir prods")
        for prod in Prods:
            print (prod)
            metros= metros + prod.metroslineales
        for prod in Prods:
            if (metros != 0 and prod.formato != 0):
                trimpond=trimpond+( (int(prod.trim)/int(prod.formato))*prod.metroslineales )/metros
            else:
                trimpond=0

        MLprodreal.append(metros)
        trims.append(trimpond)

    #detalles = DetalleProg.objects.filter(programma=orden)
    #meses=Meses.objects.all().order_by('mesnum')
    #orderinfos= OrderInfo.objects.all().order_by('SO')

    print("carga datos completada")

    return render(request, template_name, {'producciones':producciones,'maquinas': maquinas, 'dias': dias, 'MLprodreal': MLprodreal, 'trims': trims } )# ,'orderinfos':orderinfos, 'maquinas': maquinas, 'meses':meses, 'detalles':detalles})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?



def kpi_conv(request):
    template_name = 'blog/kpi_conv.html'
    ordenes = OrdenProg.objects.all().order_by('fecha_programa') #filtro para que sólo considere las órdenes de apartir de cierta fecha, para que no demore tanto la carga. #Post.objects.filter(published_date__isnull=True).order_by('create_date')
    turnos = Turnos.objects.all()
    maquinas = Maquinas.objects.all()
    reales = ProdReal.objects.all()
    padrones=Padron.objects.all()
    orderinfos = OrderInfo.objects.all()
    #detallesProg = DetalleProg.objects.all()

    ini= ordenes[0].fecha_programa
    fin= ordenes.reverse()[0].fecha_programa

    detallefilt = DetalleProg.objects.filter(datefin__gte=ini, datefin__lt=fin).order_by('datefin')


    ini= ordenes[0].fecha_programa.replace(hour= 0, minute=0, second=0, microsecond=0)#-timedelta(days=1)
    fin= ordenes.reverse()[0].fecha_programa.replace(hour= 0, minute=0, second=0, microsecond=0)

    auxfecha=ini

    auxconteoFFG=0
    auxconteoFFW=0
    auxconteoTCY=0
    auxconteoHCR=0
    auxconteoWRD=0
    auxconteoDRO=0

    auxm2progFFG=0
    auxm2progFFW=0
    auxm2progTCY=0
    auxm2progHCR=0
    auxm2progWRD=0
    auxm2progDRO=0

    aux2conteoFFG=0
    aux2conteoFFW=0
    aux2conteoTCY=0
    aux2conteoHCR=0
    aux2conteoWRD=0
    aux2conteoDRO=0

    auxm2prodFFG=0
    auxm2prodFFW=0
    auxm2prodTCY=0
    auxm2prodHCR=0
    auxm2prodWRD=0
    auxm2prodDRO=0

    auxm2prodtotFFG=0
    auxm2prodtotFFW=0
    auxm2prodtotTCY=0
    auxm2prodtotHCR=0
    auxm2prodtotWRD=0
    auxm2prodtotDRO=0


    while True:
        #print(auxfecha)
        #aux=fecha actual

        # para cada día = aux. Paso por todos los detalleprogs en ese rango, escaneo por turno y máquina. Cuento todas las ID que existen y las agrego al compID.
        #
        #ordenes = OrdeneProg.objects.filter(datefin__gte=orden
        for turno in turnos:
            for maquina in maquinas:
                for orden in ordenes:

                    flag=0
                    try:
                        if orden.get_next_by_fecha_programa()!=None:
                            flag=1 #flag1= si hay programa posterior. flag2=si se encontró el padrón en el maestropadrón
                            #print("con next by fecha")
                    except orden.DoesNotExist:
                        flag=0
                        #print("sin next by fecha")

                    if flag==1:#flag1= si hay programa posterior. flag2=si se encontró el padrón en el maestropadrón
                        #print("orden: " + str(orden.fecha_programa))
                        ordennext=orden.get_next_by_fecha_programa()
                        #print("next: " + str(ordennext.fecha_programa))
                        detallefilt = DetalleProg.objects.filter(programma= orden, datefin__gte=orden.fecha_programa, datefin__lt=ordennext.fecha_programa, datefinajustada=auxfecha, workcenter = maquina.maquina, turno=turno.turno).order_by('datefin')
                        #organizo los datos por días ajustados
                        #print("días ajustados en período: ")
                        for detalle in detallefilt:

                            ###################

                            if detalle.workcenter=="FFG":
                                #print("id: " + str(detalle.orderId))
                                auxconteoFFG=auxconteoFFG+1 # se suma a lo Programado

                                #print("calculo la cantidad de m2 programadas en este ")
                                flag2=0#flag= si hay programa posterior. flag2=si se encontró el padrón en el maestropadrón
                                try:
                                    orderinf= OrderInfo.objects.filter(orderId=detalle.orderId)[0]
                                    padron=padrones.filter(padron=orderinf.padron)[0]
                                    flag2=1

                                except:
                                    print("padrón no encontrado")


                                if flag2==1:
                                    m2programado= (orderinf.qOrd * padron.m2uni) #* padron.uxg) #cuando arregle el qIn de la tabla de aql puedo usar esto
                                    auxm2progFFG = auxm2progFFG + m2programado #por ahora lo puedo sacar del orderinfo auqnue no es confiable.
                                    print("padron: " + str(padron) + " m2 programados: " + str(m2programado))





                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoFFG=aux2conteoFFG+1
                                    if flag2==1:#Lo que se produjo dentro del turno programado
                                        real = ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter)[0]
                                        auxm2prodFFG=auxm2prodFFG+ (real.qProd * padron.m2uni)


                            ################33


                            elif detalle.workcenter=="FFW":
                                #print("id: " + str(detalle.orderId))
                                auxconteoFFW=auxconteoFFW+1 # se suma a lo Programado

                                #print("calculo la cantidad de m2 programadas en este ")
                                flag2=0#flag= si hay programa posterior. flag2=si se encontró el padrón en el maestropadrón
                                try:
                                    orderinf= OrderInfo.objects.filter(orderId=detalle.orderId)[0]
                                    padron=padrones.filter(padron=orderinf.padron)[0]
                                    flag2=1

                                except:
                                    print("padrón no encontrado")


                                if flag2==1:
                                    m2programado= (orderinf.qOrd * padron.m2uni) #* padron.uxg) #cuando arregle el qIn de la tabla de aql puedo usar esto
                                    auxm2progFFW = auxm2progFFW + m2programado #por ahora lo puedo sacar del orderinfo auqnue no es confiable.
                                    print("padron: " + str(padron) + " m2 programados: " + str(m2programado))





                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoFFW=aux2conteoFFW+1
                                    if flag2==1:#Lo que se produjo dentro del turno programado
                                        real = ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter)[0]
                                        auxm2prodFFW=auxm2prodFFW+ (real.qProd * padron.m2uni)



                            ###########

                            elif detalle.workcenter=="TCY":
                                #print("id: " + str(detalle.orderId))
                                auxconteoTCY=auxconteoTCY+1 # se suma a lo Programado

                                #print("calculo la cantidad de m2 programadas en este ")
                                flag2=0#flag= si hay programa posterior. flag2=si se encontró el padrón en el maestropadrón
                                try:
                                    orderinf= OrderInfo.objects.filter(orderId=detalle.orderId)[0]
                                    padron=padrones.filter(padron=orderinf.padron)[0]
                                    flag2=1

                                except:
                                    print("padrón no encontrado")


                                if flag2==1:
                                    m2programado= (orderinf.qOrd * padron.m2uni) #* padron.uxg) #cuando arregle el qIn de la tabla de aql puedo usar esto
                                    auxm2progTCY = auxm2progTCY + m2programado #por ahora lo puedo sacar del orderinfo auqnue no es confiable.
                                    print("padron: " + str(padron) + " m2 programados: " + str(m2programado))





                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoTCY=aux2conteoTCY+1
                                    if flag2==1:#Lo que se produjo dentro del turno programado
                                        real = ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter)[0]
                                        auxm2prodTCY=auxm2prodTCY+ (real.qProd * padron.m2uni)

                            ##########

                            elif detalle.workcenter=="HCR":
                                #print("id: " + str(detalle.orderId))
                                auxconteoHCR=auxconteoHCR+1
                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoHCR=aux2conteoHCR+1

                            elif detalle.workcenter=="WRD":
                                #print("id: " + str(detalle.orderId))
                                auxconteoWRD=auxconteoWRD+1
                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoWRD=aux2conteoWRD+1

                            elif detalle.workcenter=="DRO":
                                #print("id: " + str(detalle.orderId))
                                auxconteoDRO=auxconteoDRO+1
                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoDRO=aux2conteoDRO+1

                        #for detalle in detallefilt:
                        #print( "  " + str(detalle) + "  " + str(detalle.datefin))
                    else:
                        #print("orden: " + str(orden.fecha_programa))
                        #print("error, no hay next")
                        detallefilt = DetalleProg.objects.filter(programma= orden, datefin__gte=orden.fecha_programa, datefinajustada=auxfecha, workcenter = maquina.maquina, turno=turno.turno).order_by('datefin')
                        for detalle in detallefilt:

                            if detalle.workcenter=="FFG":
                                auxconteoFFG=auxconteoFFG+1
                                #print("id: " + str(detalle.orderId))
                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoFFG=aux2conteoFFG+1

                            elif detalle.workcenter=="FFW":
                                auxconteoFFW=auxconteoFFW+1
                                #print("id: " + str(detalle.orderId))
                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoFFW=aux2conteoFFW+1

                            elif detalle.workcenter=="TCY":
                                auxconteoTCY=auxconteoTCY+1
                                #print("id: " + str(detalle.orderId))
                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoTCY=aux2conteoTCY+1

                            elif detalle.workcenter=="HCR":
                                auxconteoHCR=auxconteoHCR+1
                                #print("id: " + str(detalle.orderId))
                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoHCR=aux2conteoHCR+1

                            elif detalle.workcenter=="WRD":
                                #print("id: " + str(detalle.orderId))
                                auxconteoWRD=auxconteoWRD+1
                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoWRD=aux2conteoWRD+1

                            elif detalle.workcenter=="DRO":
                                #print("id: " + str(detalle.orderId))
                                auxconteoDRO=auxconteoDRO+1
                                if ProdReal.objects.filter(orderId=detalle.orderId, datefinajustada=detalle.datefinajustada, turno=detalle.turno, maquina=detalle.workcenter).exists():
                                    aux2conteoDRO=aux2conteoDRO+1

                        #for detalle in detallefilt:
                        #print( "  " + str(detalle) + "  " + str(detalle.datefin))


            #print("Fecha: " + str(auxfecha))
            #print("turno: " + str(turno))

                    #print("N° órdenes programadas: "+ str(auxconteo))

            dia, created =DiaConv2.objects.get_or_create(diaajust=auxfecha, turno=turno)

            dia.semana=auxfecha.isocalendar()[1]
            dia.mes=auxfecha.month
            dia.anno=auxfecha.isocalendar()[0]

            dia.prodIdFFG=aux2conteoFFG #lo que se produjo dentro del turno
            dia.progIdFFG=auxconteoFFG

            dia.prodM2FFG=auxm2prodFFG
            dia.progM2FFG=auxm2progFFG

            dia.prodIdFFW=aux2conteoFFW
            dia.progIdFFW=auxconteoFFW

            dia.prodIdTCY=aux2conteoTCY
            dia.progIdTCY=auxconteoTCY

            dia.prodIdHCR=aux2conteoHCR
            dia.progIdHCR=auxconteoHCR

            dia.prodIdWRD=aux2conteoWRD
            dia.progIdWRD=auxconteoWRD

            dia.prodIdDRO=aux2conteoDRO
            dia.progIdDRO=auxconteoDRO
         #dia.compId=auxconteo

         #Cuento las producciones reales que se hicieron en ese turno, independiente de si estaba programado o no

            dia.prod2IdFFG=ProdReal.objects.filter(datefinajustada=auxfecha, turno=turno, maquina="FFG").count()#Lo que se produjo dentro y fuera del turno
            dia.prod2IdFFW=ProdReal.objects.filter(datefinajustada=auxfecha, turno=turno, maquina="FFW").count()
            dia.prod2IdTCY=ProdReal.objects.filter(datefinajustada=auxfecha, turno=turno, maquina="TCY").count()
            dia.prod2IdHCR=ProdReal.objects.filter(datefinajustada=auxfecha, turno=turno, maquina="HCR").count()
            dia.prod2IdWRD=ProdReal.objects.filter(datefinajustada=auxfecha, turno=turno, maquina="WRD").count()
            dia.prod2IdDRO=ProdReal.objects.filter(datefinajustada=auxfecha, turno=turno, maquina="DRO").count()



            for maquina in maquinas:
                for real in ProdReal.objects.filter(datefinajustada=auxfecha, turno=turno, maquina=maquina):
                    #busco el padrón de cada prog real para sacarle los m2
                    flag2=0
                    try:
                        orderinf= OrderInfo.objects.filter(orderId=real.orderId)[0]
                        padron=Padron.objects.filter(padron=orderinf.padron)[0]
                        flag2=1
                    except:
                        print("padrón no encontrado")

                    if flag2==1:
                        m2producido= (real.qProd * padron.m2uni) #cuando arregle el qIn de la tabla de aql puedo usar esto
                        print("m2 producido tot: " + str(m2producido))
                        #auxm2progFFG = m2programado #por ahora lo puedo sacar del orderinfo auqnue no es confiable.
                        if str(maquina)=="FFG":
                            auxm2prodtotFFG=auxm2prodtotFFG + m2producido
                        if str(maquina)=="FFW":
                            auxm2prodtotFFW=auxm2prodtotFFW + m2producido



            dia.prod2M2FFG=auxm2prodtotFFG
            dia.prod2M2FFW=auxm2prodtotFFW
            dia.prod2M2TCY=auxm2prodtotTCY
            dia.prod2M2WRD=auxm2prodtotWRD
            dia.prod2M2HCR=auxm2prodtotHCR
            dia.prod2M2DRO=auxm2prodtotDRO





            dia.save()

            auxconteoFFG=0
            auxconteoFFW=0
            auxconteoTCY=0
            auxconteoHCR=0
            auxconteoWRD=0
            auxconteoDRO=0

            auxm2progFFG=0
            auxm2progFFW=0
            auxm2progTCY=0
            auxm2progHCR=0
            auxm2progWRD=0
            auxm2progDRO=0

            aux2conteoFFG=0
            aux2conteoFFW=0
            aux2conteoTCY=0
            aux2conteoHCR=0
            aux2conteoWRD=0
            aux2conteoDRO=0

            auxm2prodFFG=0
            auxm2prodFFW=0
            auxm2prodTCY=0
            auxm2prodHCR=0
            auxm2prodWRD=0
            auxm2prodDRO=0

            auxm2prodtotFFG=0
            auxm2prodtotFFW=0
            auxm2prodtotTCY=0
            auxm2prodtotHCR=0
            auxm2prodtotWRD=0
            auxm2prodtotDRO=0


        auxfecha =auxfecha + timedelta(days=1)




        if(auxfecha>fin):
            break

    diaconvs= DiaConv2.objects.all().order_by('diaajust')
    return render(request, template_name, {'ordenes':ordenes, 'diaconvs':diaconvs})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?


def res_pareados(request):

    print("carga datos iniciada")
    template_name = 'blog/res_pareados.html'
    orden = OrdenProg.objects.latest('fecha_programa') #Post.objects.filter(published_date__isnull=True).order_by('create_date')
    turnos = Turnos.objects.all()
    maquinas = Maquinas.objects.all()
    detalles = DetalleProg.objects.filter(programma=orden)
    meses=Meses.objects.all().order_by('mesnum')
    orderinfos= OrderInfo.objects.all().order_by('SO')

    print("carga datos completada")

    return render(request, template_name, {'orden':orden, 'orderinfos':orderinfos, 'maquinas': maquinas, 'meses':meses, 'detalles':detalles})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?




def kpi_conv_res(request):
    template_name = 'blog/kpi_conv.html'
    ordenes = OrdenProg.objects.all().order_by('fecha_programa') #Post.objects.filter(published_date__isnull=True).order_by('create_date')
    turnos = Turnos.objects.all()
    maquinas = Maquinas.objects.all()
    reales = ProdReal.objects.all()
    meses=Meses.objects.filter( Q(mesnum=datetime.now().month-1) | Q(mesnum=datetime.now().month-2) | Q(mesnum=datetime.now().month) ).order_by('mesnum')#Filtro el mes pasado y el actual.
    semanas=Semanas.objects.all().order_by('-semana')

    diaconvs= DiaConv2.objects.filter(mes__gte=datetime.now().month-2).order_by('-diaajust')
    return render(request, template_name, {'ordenes':ordenes, 'diaconvs':diaconvs, 'maquinas': maquinas, 'meses':meses, 'reales':reales, 'semanas':semanas})#     , "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?





def carga_maestrop(request):
    template_name = 'blog/carga_maestrop.html'
    prodsreales = ProdReal.objects.all()
    if request.method == "POST":

        form = PruebaModForm(request.POST) ##Ojo esta sí sirve, es el Ultrafile donde se pega el excel
        if form.is_valid():
            datocrudo=form.cleaned_data["ultrafile"]
            #print("datocrudo.clean: " + datocrudo)
        else:
            datocrudo=form.data["ultrafile"]
            datoprocesado=datocrudo.split(",;,")
            #print("datoprocesado1:")
            #print(datoprocesado)
            for i in range (len(datoprocesado)):
                datoprocesado[i]=datoprocesado[i].split(",")
            #print(datoprocesado)
            ####### identifica las columnas y crea los objetos que me interesanself.
            colPadron = None
            colUxg = None
            colM2uni = None

            colStatus = None


            for i in range(len(datoprocesado[0])):
                if datoprocesado[0][i] == "Articulo Item":
                    #print("columna creación en: " + str(i))
                    colPadron = i

                if datoprocesado[0][i] == "Unid. x Golpe":
                    #print("columna transaction en: " + str(i))
                    colUxg = i

                if datoprocesado[0][i] == "M2 unitario":
                    #print("columna transaction en: " + str(i))
                    colM2uni = i

                if datoprocesado[0][i] == "STATUS":
                    #print("columna transaction en: " + str(i))
                    colStatus= i

            #fecha_now = datetime.now()
            #nuevacarga=CargaProducciones.objects.get_or_create(fecha_carga=fecha_now)
            '''
            #ejemplo

            obj, created = Person.objects.get_or_create(
                first_name='John',
                last_name='Lennon',
                defaults={'birthday': date(1940, 10, 9)},
            )
            '''


            for i in range(1,len(datoprocesado)):
                #if i==1:
                #    Reciencreado=ProdReal.objects.get_or_create(cliente=datoprocesado[i][colCliente], orderId=datoprocesado[i][colOrderId], orderIdPrev="Primero", orderIdPost=datoprocesado[i+1][colOrderId])
                #elif i==len(datoprocesado):
                #    Reciencreado=ProdReal.objects.get_or_create(cliente=datoprocesado[i][colCliente], orderId=datoprocesado[i][colOrderId], orderIdPrev=datoprocesado[i-1][colOrderId], orderIdPost=datoprocesado[i+1][colOrderId])
                #else:
                #datefinajustada_datetime = datetime.strptime(datoprocesado[i][colDatefinajust], "%d-%m-%Y")#"%d-%m-%Y %H:%M"
                #datefin_datetime = datetime.strptime(datoprocesado[i][colDatefin], "%d-%m-%Y %H:%M:%S")

                try:
                    if (datoprocesado[i][colStatus] != "NULO"):
                        Padron.objects.get_or_create(padron=datoprocesado[i][colPadron], m2uni=datoprocesado[i][colM2uni], uxg=int(datoprocesado[i][colUxg]), status=datoprocesado[i][colStatus])
                #################
                except:
                    print("Error en padrón: "+ datoprocesado[i][colPadron])

    else:
        form = PruebaModForm()

    return render(request, template_name, {'form':form})





def minuta_detail(request, pk):
    minuta = get_object_or_404(Minuta, pk=pk)
    return render(request, 'blog/minuta_detail.html', {'minuta': minuta})

def minutas(request):
    minutas = Minuta.objects.all()
    template_name = 'blog/minutas.html'
    return render(request, template_name, {'minutas':minutas})


def minuta_new(request):

    if request.method == "POST":
        form = MinutaForm(request.POST)
        if form.is_valid():
            minuta = form.save(commit=False)
            #minuta.author = request.user
            #minuta.published_date = timezone.now()
            minuta.save()
            return redirect('minuta_detail', pk=minuta.pk)
    else:
        form = MinutaForm()
    return render(request, 'blog/minuta_edit.html', {'form': form})
#####

class OrdenProgCorrDetailView(DetailView):
    context_object_name = 'ordenprogcorr'
    model = OrdenProgCorr

    def actualizadatos(self, pk):
        #referencia = OrdenProg.objects.filter(pk=pk[0])#ojo tmabién puede ser con el get. ahí hay que poner el [0] ya que no acepta más de un resultado
        referencia = OrdenProgCorr.objects.get(pk=pk)
        detalles = DetalleProgCorr.objects.filter(programma = referencia)



    def get_context_data(self, **kwargs):

        pk = self.kwargs['pk']# this is the primary key from your URL
        # your other code

        print(pk)

        self.actualizadatos(pk)

        context = super().get_context_data(**kwargs)
        context['detalleprog'] = DetalleProgCorr.objects.filter(programma = OrdenProgCorr.objects.get(pk=pk), datefinajustada__lte = OrdenProgCorr.objects.get(pk=pk).horizontefin ).order_by('datefin')#.filter(published_date__isnull=True).order_by('-published_date')
        context['prodreal'] = ProdRealCorr.objects.filter(datefin__gte=OrdenProgCorr.objects.get(pk=pk).fecha_programa).order_by('datefin')#Acá hay que filtrar para que sean las órdenes reales desde la fecha del programa de referencia en adelante
        #context['maquinas'] = Maquinas.objects.all()#.filter(published_date__isnull=True).order_by('-published_date')
        #context['turnos'] = Turnos.objects.all()#.filter(published_date__isnull=True).order_by('-published_date')
        #context['orderinfos'] = OrderInfo.objects.all() #filtrar para que sólo mande los que están dentro del detalleprog?
        #context['padrones'] = Padron.objects.all()#Agregarle al orderinfo fecha de subida y filtrar para que busque?
        #context['maquinas'] = maquinas
        return context


####

#####

class OrdenProgDetailView(DetailView):
    context_object_name = 'ordenprog'
    model = OrdenProg

    def actualizadatos(self, pk):
        #referencia = OrdenProg.objects.filter(pk=pk[0])#ojo tmabién puede ser con el get. ahí hay que poner el [0] ya que no acepta más de un resultado
        referencia = OrdenProg.objects.get(pk=pk)
        detalles = DetalleProg.objects.filter(programma = referencia)
        ordenprevia = OrdenProg.objects.get(pk=pk).get_previous_by_fecha_programa()



    def get_context_data(self, **kwargs):

        pk = self.kwargs['pk']# this is the primary key from your URL
        # your other code

        print(pk)

        #self.actualizadatos(pk)

        context = super().get_context_data(**kwargs)
        detalleprog2 = DetalleProg.objects.filter(programma = OrdenProg.objects.get(pk=pk), datefinajustada__lte = OrdenProg.objects.get(pk=pk).horizontefin )#.filter(published_date__isnull=True).order_by('-published_date')

        detalleprogprev = DetalleProg.objects.filter(programma = OrdenProg.objects.get(pk=pk).get_previous_by_fecha_programa(), datefin__lt= OrdenProg.objects.get(pk=pk).fecha_programa, datefinajustada__gte = OrdenProg.objects.get(pk=pk).fecha_programa.date() )

        context['detalleprog'] = detalleprog2.union(detalleprogprev, all=True).order_by('datefin')

        context['prodreal'] = ProdReal.objects.filter(datefin__gte=OrdenProg.objects.get(pk=pk).fecha_programa + timedelta(days=-1), datefin__lt=OrdenProg.objects.get(pk=pk).horizontefin + timedelta(days=1)).order_by('datefin')#Acá hay que filtrar para que sean las órdenes reales desde la fecha del programa de referencia en adelante
        context['maquinas'] = Maquinas.objects.all()#.filter(published_date__isnull=True).order_by('-published_date')
        context['turnos'] = Turnos.objects.all()#.filter(published_date__isnull=True).order_by('-published_date')
        context['orderinfos'] = OrderInfo.objects.all() #filtrar para que sólo mande los que están dentro del detalleprog?
        #context['padrones'] = Padron.objects.all()#Agregarle al orderinfo fecha de subida y filtrar para que busque?

        return context


####


class Inicio(View): #ESto se puede cambiar a format based view (para que no sea clase)
    'blog/inicio.html'
    #suma el total de órdenes producidas reales para los últimos 5 días.
    datefinajustadaini = datetime(2018, 9 ,12,0,0)
    datefinajustadafin=  datetime(2018, 9 ,25,0,0)
    datefinajustadaobj=  datetime(1,1,1,0,0)
    Producciones = []
    Dias=[]
    datefinajustadaobj = datefinajustadaini


    while True:
        Dias.append(datefinajustadaobj)
        datefinajustadaobj=datefinajustadaobj + timedelta(days=1)
        if datefinajustadaobj > datefinajustadafin :
            break



    def get(self, request, *args, **kwargs):
        return render(request, 'blog/inicio.html', {"customers": 111})


#Esta es la que les manda el dato al gráfico de inicio.
def get_data_inicio(request, *args, **kwargs):
    orden = OrdenProg.objects.all().order_by('-pk').first()
    detallesProg = DetalleProg.objects.filter(programma= orden)
    prodsreales = ProdReal.objects.all()
    padrones = Padron.objects.all()

    datefinajustadaini = datetime.now().replace(hour= 0, minute=0, second=0, microsecond=0) - timedelta(days=5)#datetime(2018, 9 ,12,0,0)
    datefinajustadafin=  datetime.now().replace(hour= 0, minute=0, second=0, microsecond=0) + timedelta(days=4)#datetime(2018, 9 ,25,0,0)
    datefinajustadaobj=  datetime(1,1,1,0,0)
    Producciones = []
    M2RDC=[]
    M2Flexo=[]
    M2ProgFlexo=[]
    M2ProgRDC=[]
    Me2=[]
    Dias=[]
    Dias2=[]
    datefinajustadaobj = datefinajustadaini
    #datefinajustadaobj2 = datefinajustadaini.strftime("%d %m %Y")


    while True:
        Dias.append(datefinajustadaobj)
        Dias2.append(datefinajustadaobj.strftime("%d %m %Y"))
        datefinajustadaobj=(datefinajustadaobj + timedelta(days=1))
        #datefinajustadaobj2=(datefinajustadaobj + timedelta(days=1)).strftime("%d %m %Y")
        if datefinajustadaobj >= datefinajustadafin :
            break


    print("ahora agrego el número de producciones que se hicieron en cada día")
    for dia in Dias:
        #Producciones.append(ProdReal.objects.filter(datefinajustada=dia).count())
        #Sumo todos los m2 de cada orden producida en el día
        sumaM2RDC=0
        sumaM2Flexo=0
        sumaProgFlexo=0
        sumaProgRDC=0
        auxsum=0
        for prod in ProdReal.objects.filter(datefinajustada=dia):
            auxsum=0
            print("padron encontrado: " + prod.padron)
            for pad in Padron.objects.filter(padron=prod.padron):
                #print(pad.m2uni)
                auxsum=pad.m2uni

            #pad = get_object_or_404(Padron, padron=prod.padron)
            if (prod.maquina == "FFG" or prod.maquina == "FFW" or prod.maquina == "TCY"):
                sumaM2Flexo= sumaM2Flexo + (int(prod.qProd)*auxsum)/1000
            if (prod.maquina == "WRD" or prod.maquina == "HCR" or prod.maquina == "DRO"):
                sumaM2RDC= sumaM2RDC + (int(prod.qProd)*auxsum)/1000


        for detalle in DetalleProg.objects.filter(datefinajustada=dia, programma= orden):
            auxsum=0
            flag=0
            print("buscando m2 de padrón progrmado en Id:" + str(detalle.orderId))

            #detallefilt = DetalleProg.objects.filter(programma= orden, datefin__gte=orden.fecha_programa, datefin__lt=ordennext.fecha_programa, datefinajustada=auxfecha, workcenter = maquina.maquina, turno=turno.turno).order_by('datefin')
            try:
                print("Orderinfo encontrado: " + str(OrderInfo.objects.filter(orderId=detalle.orderId)[0]))

            except:
                print("Orderinfo no encontrado")

            try:
                orderinf= OrderInfo.objects.filter(orderId=detalle.orderId)[0]
                print("padron encontrado: " + str(Padron.objects.filter(padron=orderinf.padron)[0]))
                flag=1
            except:
                print("Padrón no encontrado")

            if flag==1:

                    orderinf= OrderInfo.objects.filter(orderId=detalle.orderId)[0]
                    padron=Padron.objects.filter(padron=orderinf.padron)[0]
                    auxsum= (orderinf.qOrd * padron.m2uni)
                    print("auxsum: " + str(auxsum))

            if (detalle.workcenter == "FFG" or detalle.workcenter == "FFW" or detalle.workcenter == "TCY"):
                sumaProgFlexo= sumaProgFlexo + (auxsum)/1000
            if (detalle.workcenter == "WRD" or detalle.workcenter == "HCR" or detalle.workcenter == "DRO"):
                sumaProgRDC= sumaProgRDC + (auxsum)/1000





        M2Flexo.append(round(sumaM2Flexo,0))
        M2RDC.append(round(sumaM2RDC,0))
        M2ProgFlexo.append(round(sumaProgFlexo,0))
        M2ProgRDC.append(round(sumaProgRDC,0))

        #Sumo los metros programado por cada día según el último programaself.




    print(Producciones)

    for dia2 in Dias2:
        print(dia2)



    labels=["red","blue","green"]
    default_items=[123, 124, 432]
    data = {
    "labels": labels,
    "default": default_items,
    "labelDias": Dias2,
    "M2Flexo":M2Flexo,
    "M2RDC":M2RDC,
    "M2ProgFlexo":M2ProgFlexo,
    "M2ProgRDC":M2ProgRDC,
    }
    print("Enviando Json Datos Graph")
    return JsonResponse(data)#http response con el datatype de JS



def res_corr(request):
    template_name = 'blog/resumencorr.html'
    ordenescorr = OrdenProgCorr.objects.all().order_by('-fecha_programa') #Post.objects.filter(published_date__isnull=True).order_by('create_date')
    detallesProgcorr = DetalleProgCorr.objects.all()

    return render(request, template_name, {'ordenescorr':ordenescorr, "detallesProgcorr": detallesProgcorr})#acá le puedo decir que los mande ordenados por fecha?




def res_conv_v2(request):
    template_name = 'blog/resumenconv.html'
    ordenes = OrdenProg.objects.all().order_by('-fecha_programa')[:6] #Post.objects.filter(published_date__isnull=True).order_by('create_date')
    detallesProg = DetalleProg.objects.all()

    return render(request, template_name, {'ordenes':ordenes, "detallesProg": detallesProg})#acá le puedo decir que los mande ordenados por fecha?

class ResConvView(ListView):
    model = OrdenProg


def carga_prod_real(request):
    template_name = 'blog/cargaprodreal.html'
    prodsreales = ProdReal.objects.all()
    if request.method == "POST":

        form = PruebaModForm(request.POST) ##Ojo esta sí sirve, es el Ultrafile donde se pega el excel
        if form.is_valid():
            datocrudo=form.cleaned_data["ultrafile"]
            print("datocrudo.clean: " + datocrudo)
        else:
            datocrudo=form.data["ultrafile"]
            datoprocesado=datocrudo.split(",;,")
            print("datoprocesado1:")
            print(datoprocesado)
            for i in range (len(datoprocesado)):
                datoprocesado[i]=datoprocesado[i].split(",")
            print(datoprocesado)
            ####### identifica las columnas y crea los objetos que me interesanself.
            colCliente = None
            colOrderId = None
            colPadron = None
            colDatefin = None
            colDatefinajust = None
            colTurno = None
            colUnisprod = None
            colMaquina = None
            colqOrd = None

            for i in range(len(datoprocesado[0])):
                if datoprocesado[0][i] == "Cliente":
                    #print("columna creación en: " + str(i))
                    colCliente = i

                if datoprocesado[0][i] == "No. Orden":
                    #print("columna transaction en: " + str(i))
                    colOrderId = i

                if datoprocesado[0][i] == "ID Especificacion":
                    #print("columna transaction en: " + str(i))
                    colPadron = i

                if datoprocesado[0][i] == "Fecha Hora de termino":
                    #print("columna transaction en: " + str(i))
                    colDatefin = i

                if datoprocesado[0][i] == "Fecha termino ajustada":
                    #print("columna transaction en: " + str(i))
                    colDatefinajust = i


                if datoprocesado[0][i] == "Turno":
                    #print("columna transaction en: " + str(i))
                    colTurno = i

                if datoprocesado[0][i] == "Cantidad Ordenada":#Esto en verdad corresponde a las cajas producidas
                    #print("columna transaction en: " + str(i))
                    colqOrd = i


                if datoprocesado[0][i] == "Laminas Producidas":#Esto en verdad corresponde a las cajas producidas
                    #print("columna transaction en: " + str(i))
                    colUnisprod = i


                if datoprocesado[0][i] == "Maquina":
                    #print("columna transaction en: " + str(i))
                    colMaquina= i
            #fecha_now = datetime.now()
            #nuevacarga=CargaProducciones.objects.get_or_create(fecha_carga=fecha_now)
            '''
            #ejemplo

            obj, created = Person.objects.get_or_create(
                first_name='John',
                last_name='Lennon',
                defaults={'birthday': date(1940, 10, 9)},
            )
            '''


            for i in range(1,len(datoprocesado)):
                #if i==1:
                #    Reciencreado=ProdReal.objects.get_or_create(cliente=datoprocesado[i][colCliente], orderId=datoprocesado[i][colOrderId], orderIdPrev="Primero", orderIdPost=datoprocesado[i+1][colOrderId])
                #elif i==len(datoprocesado):
                #    Reciencreado=ProdReal.objects.get_or_create(cliente=datoprocesado[i][colCliente], orderId=datoprocesado[i][colOrderId], orderIdPrev=datoprocesado[i-1][colOrderId], orderIdPost=datoprocesado[i+1][colOrderId])
                #else:
                datefinajustada_datetime = datetime.strptime(datoprocesado[i][colDatefinajust], "%d-%m-%Y")#"%d-%m-%Y %H:%M"
                datefin_datetime = datetime.strptime(datoprocesado[i][colDatefin], "%d-%m-%Y %H:%M:%S")

                try:
                    o = ProdReal.objects.filter(cliente=datoprocesado[i][colCliente], padron=datoprocesado[i][colPadron], orderId=datoprocesado[i][colOrderId], datefin=datefin_datetime, datefinajustada=datefinajustada_datetime, turno=datoprocesado[i][colTurno], maquina=datoprocesado[i][colMaquina])[0]
                except:
                    o, created = ProdReal.objects.get_or_create(cliente=datoprocesado[i][colCliente], padron=datoprocesado[i][colPadron], orderId=datoprocesado[i][colOrderId], orderIdPrev="pendiente", orderIdPost="Final", datefin=datefin_datetime, datefinajustada=datefinajustada_datetime, turno=datoprocesado[i][colTurno], maquina=datoprocesado[i][colMaquina])


                o.qProd=datoprocesado[i][colUnisprod]
                o.qOrd=datoprocesado[i][colqOrd]
                o.semana=datefin_datetime.isocalendar()[1]
                o.mes=datefin_datetime.month
                o.anno=datefin_datetime.isocalendar()[0]

                o.save()


            #################


    else:
        form = PruebaModForm()

    return render(request, template_name, {'form':form})



#################################################


''' ### Referencia de caraorderinfo que no se cae.


def carga_orderinfo(request):
    pruebamods = PruebaMod.objects.all()

    template_name = 'blog/cargaorderinfo.html'

    if request.method == "POST":
        form = PruebaModForm(request.POST)
        if form.is_valid():


            datocrudo=form.cleaned_data["ultrafile"]
            datoprocesado=datocrudo.split(";")

            #PruebaTabla.objects.create(item1=datoprocesado[0],item2=datoprocesado[1],item3=datoprocesado[2])

            post = form.save(commit=False)

            post.save()

        else:
            datocrudo=form.data["ultrafile"]

            datoprocesado=datocrudo.split(",;,")

            for i in range (len(datoprocesado)):
                datoprocesado[i]=datoprocesado[i].split(",")

            ####### identifica las columnas y crea los objetos que me interesanself.
            colOrderId = None#Esta la tengo que pasar a datetime
            colPadron = None
            colCliente = None#Esta la tengo que pasar a datetime


            for i in range(len(datoprocesado[0])):
                if datoprocesado[0][i] == "ORDERID":
                    #print("columna creación en: " + str(i))
                    colOrderId = i

                if datoprocesado[0][i] == "INTERNALSPECID":
                    #print("columna transaction en: " + str(i))
                    colPadron = i

                if datoprocesado[0][i] == "CUSTOMERNAME":
                    #print("columna Horizonteini en: " + str(i))
                    colCliente = i

            print("guardando..")
            for i in range(1,len(datoprocesado)):


                    o, created=OrderInfo.objects.get_or_create(orderId=datoprocesado[i][colOrderId])
                    o.padron=datoprocesado[i][colPadron]
                    o.cliente=datoprocesado[i][colCliente]
                    o.save()

            print("completado!")
    else:
        form = PruebaModForm()

    return render(request, template_name, {'form':form})





'''

def carga_orderinfo(request):
    pruebamods = PruebaMod.objects.all()

    template_name = 'blog/cargaorderinfo.html'

    if request.method == "POST":
        form = PruebaModForm(request.POST)
        if form.is_valid():


            datocrudo=form.cleaned_data["ultrafile"]
            datoprocesado=datocrudo.split(";")

            #PruebaTabla.objects.create(item1=datoprocesado[0],item2=datoprocesado[1],item3=datoprocesado[2])

            post = form.save(commit=False)

            post.save()

        else:
            datocrudo=form.data["ultrafile"]

            datoprocesado=datocrudo.split(",;,")

            for i in range (len(datoprocesado)):
                datoprocesado[i]=datoprocesado[i].split(",")

            ####### identifica las columnas y crea los objetos que me interesanself.
            colOrderId = None#Esta la tengo que pasar a datetime
            colPadron = None
            colCliente = None#Esta la tengo que pasar a datetime
            colSO = None
            colSOPos = None
            colQord = None
            colBlanksReq = None
            colBlanksToCorr = None


            for i in range(len(datoprocesado[0])):
                if datoprocesado[0][i] == "ORDERID":
                    #print("columna creación en: " + str(i))
                    colOrderId = i

                if datoprocesado[0][i] == "INTERNALSPECID":
                    #print("columna transaction en: " + str(i))
                    colPadron = i

                if datoprocesado[0][i] == "CUSTOMERNAME":
                    #print("columna Horizonteini en: " + str(i))
                    colCliente = i

                if datoprocesado[0][i] == "SalesOrderNo":
                    #print("columna Horizonteini en: " + str(i))
                    colSO = i

                if datoprocesado[0][i] == "SalesOrderPositionNo":
                    #print("columna Horizonteini en: " + str(i))
                    colSOPos = i

                if datoprocesado[0][i] == "ITEMSORDERED":
                    #print("columna Horizonteini en: " + str(i))
                    colQord = i

                if datoprocesado[0][i] == "BLANKSREQUIRED":
                    #print("columna Horizonteini en: " + str(i))
                    colBlanksReq = i

                if datoprocesado[0][i] == "BLANKSTOCORR":
                    #print("columna Horizonteini en: " + str(i))
                    colBlanksToCorr = i

            print("guardando..")
            for i in range(1,len(datoprocesado)):

                    '''
                    o, created=OrderInfo.objects.get_or_create(orderId=datoprocesado[i][colOrderId])
                    o.padron=datoprocesado[i][colPadron]
                    o.cliente=datoprocesado[i][colCliente]
                    o.SO=datoprocesado[i][colSO]
                    o.SOPosition=datoprocesado[i][colSOPos]
                    o.qOrd=datoprocesado[i][colQord]
                    o.blanksrequired=datoprocesado[i][colBlanksReq]
                    o.blankstocorr=datoprocesado[i][colBlanksToCorr]
                    o.save()
                    '''
                    o, created=OrderInfo.objects.get_or_create(orderId=datoprocesado[i][colOrderId])
                    o.padron=datoprocesado[i][colPadron]
                    o.cliente=datoprocesado[i][colCliente]
                    o.save()

            print("completado!")
    else:
        form = PruebaModForm()

    return render(request, template_name, {'form':form})









###############################################33


def carga_prog_corr(request):
    pruebamods = PruebaMod.objects.all()

    template_name = 'blog/carga_prog_corr.html'

    if request.method == "POST":
        form = PruebaModForm(request.POST)
        if form.is_valid():


            datocrudo=form.cleaned_data["ultrafile"]
            datoprocesado=datocrudo.split(";")

            #PruebaTabla.objects.create(item1=datoprocesado[0],item2=datoprocesado[1],item3=datoprocesado[2])

            post = form.save(commit=False)

            post.save()

        else:
            datocrudo=form.data["ultrafile"]

            datoprocesado=datocrudo.split(",;,")

            for i in range (len(datoprocesado)):
                datoprocesado[i]=datoprocesado[i].split(",")

            ####### identifica las columnas y crea los objetos que me interesanself.





            fecha_programa_datetime=datetime.now()
            fecha_programa_horini=fecha_programa_datetime.replace(hour=0, minute=0, second=0, microsecond=0)
            fecha_programa_horfin=fecha_programa_horini + timedelta(days=1)
                        #print("fecha programa")
                        #print(fecha_programa_datetime)

            OrdenProgCorr.objects.get_or_create(fecha_programa=fecha_programa_datetime, horizonteini=fecha_programa_horini, horizontefin=fecha_programa_horfin )


            colAjuste=3
            colOnda=4
            colFormato=5
            colCarton=6
            colMl=7
            colTrim=8
            colPapeles=9
            colDatefin=12
            colDatefinajustada=19
            colTurno=20



            print("guardando..")
            for i in range(1,len(datoprocesado)):
                datefinajustada_datetime = datetime.strptime(datoprocesado[i][colDatefinajustada], "%d-%m-%Y")#"%d-%m-%Y %H:%M"
                datefin_datetime = datetime.strptime(datoprocesado[i][colDatefin], "%d-%m-%Y %H:%M")#"%d-%m-%Y %H:%M"

                DetalleProgCorr.objects.get_or_create(programma=OrdenProgCorr.objects.filter(fecha_programa=fecha_programa_datetime)[0], ajuste=datoprocesado[i][colAjuste], onda=datoprocesado[i][colOnda], formato=datoprocesado[i][colFormato], carton=datoprocesado[i][colCarton], metroslineales=datoprocesado[i][colMl], trim=datoprocesado[i][colTrim], papeles=datoprocesado[i][colPapeles],datefin=datefin_datetime, datefinajustada= datefinajustada_datetime , turno=datoprocesado[i][colTurno])
                print("completado!")
        return redirect ('res_corr')

    else:
        form = PruebaModForm()

    return render(request, template_name, {'form':form})



def carga_prod_corr(request):
    pruebamods = PruebaMod.objects.all()

    template_name = 'blog/carga_prod_corr.html'

    if request.method == "POST":
        form = PruebaModForm(request.POST)
        if form.is_valid():


            datocrudo=form.cleaned_data["ultrafile"]
            datoprocesado=datocrudo.split(";")

            #PruebaTabla.objects.create(item1=datoprocesado[0],item2=datoprocesado[1],item3=datoprocesado[2])

            post = form.save(commit=False)

            post.save()

        else:
            datocrudo=form.data["ultrafile"]

            datoprocesado=datocrudo.split(",;,")

            for i in range (len(datoprocesado)):
                datoprocesado[i]=datoprocesado[i].split(",")

            ####### identifica las columnas y crea los objetos que me interesanself.




                    #print("fecha programa")
                        #print(fecha_programa_datetime)

            #OrdenProgCorr.objects.get_or_create(fecha_programa=fecha_programa_datetime, horizonteini=fecha_programa_horini, horizontefin=fecha_programa_horfin )


            colAjuste=0
            colDatefin=1
            colDatefinajustada=11
            colOnda=5
            colFormato=6
            colCarton=4
            colMl=8
            colTrim=7
            colTurno=12



            print("guardando..")
            for i in range(1,len(datoprocesado)):
                datefinajustada_datetime = datetime.strptime(datoprocesado[i][colDatefinajustada], "%d-%m-%Y")#"%d-%m-%Y %H:%M"
                datefin_datetime = datetime.strptime(datoprocesado[i][colDatefin], "%d-%m-%Y %H:%M")#"%d-%m-%Y %H:%M"

                ProdRealCorr.objects.get_or_create(ajuste=datoprocesado[i][colAjuste], onda=datoprocesado[i][colOnda], formato=datoprocesado[i][colFormato], carton=datoprocesado[i][colCarton], metroslineales=datoprocesado[i][colMl], trim=datoprocesado[i][colTrim],datefin=datefin_datetime, datefinajustada= datefinajustada_datetime , turno=datoprocesado[i][colTurno])

                print("completado!")
        return redirect ('res_corr')

    else:
        form = PruebaModForm()

    return render(request, template_name, {'form':form})




def carga_proyeccion(request):
    pruebamods = PruebaMod.objects.all()

    template_name = 'blog/carga_proyeccion.html'

    if request.method == "POST":
        form = PruebaModForm(request.POST)
        if form.is_valid():


            datocrudo=form.cleaned_data["ultrafile"]
            datoprocesado=datocrudo.split(";")

            #PruebaTabla.objects.create(item1=datoprocesado[0],item2=datoprocesado[1],item3=datoprocesado[2])

            post = form.save(commit=False)

            post.save()

        else:
            datocrudo=form.data["ultrafile"]

            datoprocesado=datocrudo.split(",;,")

            for i in range (len(datoprocesado)):
                datoprocesado[i]=datoprocesado[i].split(",")

            ####### identifica las columnas y crea los objetos que me interesanself.





            fecha_programa_datetime=datetime.now()
            fecha_programa_horini=fecha_programa_datetime.replace(hour=0, minute=0, second=0, microsecond=0)
            fecha_programa_horfin=fecha_programa_horini + timedelta(days=1)
                        #print("fecha programa")
                        #print(fecha_programa_datetime)

            #OrdenProgCorr.objects.get_or_create(fecha_programa=fecha_programa_datetime, horizonteini=fecha_programa_horini, horizontefin=fecha_programa_horfin )


            colAjuste=3
            colOnda=4
            colFormato=5
            colCarton=6
            colMl=7
            colTrim=8
            colPapeles=9
            colDatefin=12
            colDatefinajustada=19
            colTurno=20



            print("guardando..")
            for i in range(1,len(datoprocesado)):
                datefinajustada_datetime = datetime.strptime(datoprocesado[i][colDatefinajustada], "%d-%m-%Y")#"%d-%m-%Y %H:%M"
                datefin_datetime = datetime.strptime(datoprocesado[i][colDatefin], "%d-%m-%Y %H:%M")#"%d-%m-%Y %H:%M"

                #DetalleProgCorr.objects.get_or_create(programma=OrdenProgCorr.objects.filter(fecha_programa=fecha_programa_datetime)[0], ajuste=datoprocesado[i][colAjuste], onda=datoprocesado[i][colOnda], formato=datoprocesado[i][colFormato], carton=datoprocesado[i][colCarton], metroslineales=datoprocesado[i][colMl], trim=datoprocesado[i][colTrim], papeles=datoprocesado[i][colPapeles],datefin=datefin_datetime, datefinajustada= datefinajustada_datetime , turno=datoprocesado[i][colTurno])
                print("completado!")
        return redirect ('res_corr')

    else:
        form = PruebaModForm()

    return render(request, template_name, {'form':form})




###############################################33

def resumen_inv(request): #ESto se puede cambiar a format based view (para que no sea clase)
    template_name='blog/resumeninv.html'
    #suma el total de órdenes producidas reales para los últimos 5 días.
    fotoinv = FotoInventario.objects.latest('fecha_carga')
    inventarios = FotoInventario.objects.all().order_by('-fecha_carga')[0:4]

    return render(request, template_name, {"fotoinv": fotoinv, "inventarios": inventarios })



###############################################33


def carga_inventario(request):
    pruebamods = PruebaMod.objects.all()

    template_name = 'blog/carga_inventario.html'

    if request.method == "POST":
        form = PruebaModForm(request.POST)
        if form.is_valid():


            datocrudo=form.cleaned_data["ultrafile"]
            datoprocesado=datocrudo.split(";")

            #PruebaTabla.objects.create(item1=datoprocesado[0],item2=datoprocesado[1],item3=datoprocesado[2])

            post = form.save(commit=False)

            post.save()

        else:
            datocrudo=form.data["ultrafile"]

            datoprocesado=datocrudo.split(",;,")

            for i in range (len(datoprocesado)):
                datoprocesado[i]=datoprocesado[i].split(",")

            ####### identifica las columnas y crea los objetos que me interesanself.

            colPeso=None
            colCategoria=None
            colSaldo=None
            colAntiguedad=None
            colRetenido=None

            for i in range(len(datoprocesado[0])):
                if datoprocesado[0][i] == "Peso":
                #print("columna creación en: " + str(i))
                    colPeso = i
                if datoprocesado[0][i] == "Categoria":
                #print("columna creación en: " + str(i))
                    colCategoria= i
                if datoprocesado[0][i] == "Saldo":
                #print("columna creación en: " + str(i))
                    colSaldo = i
                if datoprocesado[0][i] == "Antiguedad":
                #print("columna creación en: " + str(i))
                    colAntiguedad = i

                if datoprocesado[0][i] == "Retenido":
                #print("columna creación en: " + str(i))
                    colRetenido = i

                fecha_carga=datetime.now()
                total_kraft_kg=0
                total_blanco_kg=0
                total_CPP_kg=0
                total_otros_kg=0

                kraft_saldos_kg=0
                blanco_saldos_kg=0
                CPP_saldos_kg=0
                otros_saldos_kg=0

                kraft_saldos_un=0
                blanco_saldos_un=0
                CPP_saldos_un=0
                otros_saldos_un=0


                kraft_retenidos_un=0
                blanco_retenidos_un=0
                CPP_retenidos_un=0
                otros_retenidos_un=0

                kraft_retenidos_kg=0
                blanco_retenidos_kg=0
                CPP_retenidos_kg=0
                otros_retenidos_kg=0

                kraft_3_meses_kg=0
                blanco_3_meses_kg=0
                CPP_3_meses_kg=0
                otros_3_meses_kg=0

                kraft_6_meses_kg=0
                blanco_6_meses_kg=0
                CPP_6_meses_kg=0
                otros_6_meses_kg=0

                kraft_1_meses_kg=0
                blanco_1_meses_kg=0
                CPP_1_meses_kg=0
                otros_1_meses_kg=0


            print("calculando valores")

            for i in range(1,len(datoprocesado)):

                if datoprocesado[i][colRetenido]== "0":

                    if datoprocesado[i][colCategoria] == "Kraft":
                        total_kraft_kg = total_kraft_kg + int(datoprocesado[i][colPeso])
                        if datoprocesado[i][colSaldo] == "1":
                            kraft_saldos_kg = kraft_saldos_kg + int(datoprocesado[i][colPeso])
                            kraft_saldos_un = kraft_saldos_un + 1

                        if ( float(datoprocesado[i][colAntiguedad])>=3 and float(datoprocesado[i][colAntiguedad])<6):
                            kraft_3_meses_kg=kraft_3_meses_kg + int(datoprocesado[i][colPeso])
                        elif(float(datoprocesado[i][colAntiguedad])>=6):
                            kraft_6_meses_kg=kraft_6_meses_kg + int(datoprocesado[i][colPeso])
                        elif(float(datoprocesado[i][colAntiguedad])>=1 and float(datoprocesado[i][colAntiguedad])<3):
                            kraft_1_meses_kg=kraft_1_meses_kg + int(datoprocesado[i][colPeso])




                    elif datoprocesado[i][colCategoria] == "Blanco":
                        total_blanco_kg = total_blanco_kg + int(datoprocesado[i][colPeso])
                        if datoprocesado[i][colSaldo] == "1":
                            blanco_saldos_kg = blanco_saldos_kg + int(datoprocesado[i][colPeso])
                            blanco_saldos_un = blanco_saldos_un + 1

                        if ( float(datoprocesado[i][colAntiguedad])>=3 and float(datoprocesado[i][colAntiguedad])<6):
                            blanco_3_meses_kg=blanco_3_meses_kg + int(datoprocesado[i][colPeso])
                        elif(float(datoprocesado[i][colAntiguedad])>=6):
                            blanco_6_meses_kg=blanco_6_meses_kg + int(datoprocesado[i][colPeso])
                        elif(float(datoprocesado[i][colAntiguedad])>=1 and float(datoprocesado[i][colAntiguedad])<3):
                            blanco_1_meses_kg=blanco_1_meses_kg + int(datoprocesado[i][colPeso])


                    elif datoprocesado[i][colCategoria] == "CPP":
                        total_CPP_kg = total_CPP_kg + int(datoprocesado[i][colPeso])
                        if datoprocesado[i][colSaldo] == "1":
                            CPP_saldos_kg = CPP_saldos_kg + int(datoprocesado[i][colPeso])
                            CPP_saldos_un = CPP_saldos_un + 1

                        if ( float(datoprocesado[i][colAntiguedad])>=3 and float(datoprocesado[i][colAntiguedad])<6):
                            CPP_3_meses_kg=CPP_3_meses_kg + int(datoprocesado[i][colPeso])
                        elif(float(datoprocesado[i][colAntiguedad])>=6):
                            CPP_6_meses_kg=CPP_6_meses_kg + int(datoprocesado[i][colPeso])
                        elif(float(datoprocesado[i][colAntiguedad])>=1 and float(datoprocesado[i][colAntiguedad])<3):
                            CPP_1_meses_kg=CPP_1_meses_kg + int(datoprocesado[i][colPeso])


                    else:
                        total_otros_kg = total_otros_kg + int(datoprocesado[i][colPeso])
                        if datoprocesado[i][colSaldo] == "1":
                            otros_saldos_kg = otros_saldos_kg + int(datoprocesado[i][colPeso])
                            otros_saldos_un = otros_saldos_un + 1

                        if ( float(datoprocesado[i][colAntiguedad])>=3 and float(datoprocesado[i][colAntiguedad])<6):
                            otros_3_meses_kg=otros_3_meses_kg + int(datoprocesado[i][colPeso])
                        elif(float(datoprocesado[i][colAntiguedad])>=6):
                            otros_6_meses_kg=otros_6_meses_kg + int(datoprocesado[i][colPeso])
                        elif(float(datoprocesado[i][colAntiguedad])>=1 and float(datoprocesado[i][colAntiguedad])<3):
                            otros_1_meses_kg=otros_1_meses_kg + int(datoprocesado[i][colPeso])

                else:
                    if datoprocesado[i][colCategoria] == "Kraft":
                        kraft_retenidos_kg = kraft_retenidos_kg + int(datoprocesado[i][colPeso])
                        kraft_retenidos_un = kraft_retenidos_un + 1



                    elif datoprocesado[i][colCategoria] == "Blanco":
                        blanco_retenidos_kg = blanco_retenidos_kg + int(datoprocesado[i][colPeso])
                        blanco_retenidos_un = blanco_retenidos_un + 1


                    elif datoprocesado[i][colCategoria] == "CPP":
                        CPP_retenidos_kg = CPP_retenidos_kg + int(datoprocesado[i][colPeso])
                        CPP_retenidos_un = CPP_retenidos_un + 1


                    else:
                        otros_retenidos_kg = otros_retenidos_kg + int(datoprocesado[i][colPeso])
                        otros_retenidos_un = otros_retenidos_un + 1

            o, created=FotoInventario.objects.get_or_create(fecha_carga=fecha_carga)
            o.total_kraft_kg=total_kraft_kg
            o.total_blanco_kg=total_blanco_kg
            o.total_CPP_kg=total_CPP_kg
            o.total_otros_kg=total_otros_kg

            o.kraft_saldos_kg=kraft_saldos_kg
            o.blanco_saldos_kg=blanco_saldos_kg
            o.CPP_saldos_kg=CPP_saldos_kg
            o.otros_saldos_kg=otros_saldos_kg

            o.kraft_saldos_un=kraft_saldos_un
            o.blanco_saldos_un=blanco_saldos_un
            o.CPP_saldos_un=CPP_saldos_un
            o.otros_saldos_un=otros_saldos_un

            o.kraft_retenidos_kg=kraft_retenidos_kg
            o.blanco_retenidos_kg=blanco_retenidos_kg
            o.CPP_retenidos_kg=CPP_retenidos_kg
            o.otros_retenidos_kg=otros_retenidos_kg

            o.kraft_retenidos_un=kraft_retenidos_un
            o.blanco_retenidos_un=blanco_retenidos_un
            o.CPP_retenidos_un=CPP_retenidos_un
            o.otros_retenidos_un=otros_retenidos_un


            o.kraft_3_meses_kg=kraft_3_meses_kg
            o.blanco_3_meses_kg=blanco_3_meses_kg
            o.CPP_3_meses_kg=CPP_3_meses_kg
            o.otros_3_meses_kg=otros_3_meses_kg


            o.kraft_6_meses_kg=kraft_6_meses_kg
            o.blanco_6_meses_kg=blanco_6_meses_kg
            o.CPP_6_meses_kg=CPP_6_meses_kg
            o.otros_6_meses_kg=otros_6_meses_kg

            o.kraft_1_meses_kg=kraft_6_meses_kg
            o.blanco_1_meses_kg=blanco_6_meses_kg
            o.CPP_1_meses_kg=CPP_6_meses_kg
            o.otros_1_meses_kg=otros_6_meses_kg


            o.save()


        #return redirect ('res_inventario')

    else:
        form = PruebaModForm()

    return render(request, template_name, {'form':form})

###############################################33





def carga_prog(request):
    pruebamods = PruebaMod.objects.all()
    posts= "HOLA Q ASE"
    template_name = 'blog/cargaprog.html'

    if request.method == "POST":
        form = PruebaModForm(request.POST)
        if form.is_valid():


            datocrudo=form.cleaned_data["ultrafile"]
            datoprocesado=datocrudo.split(";")

            #PruebaTabla.objects.create(item1=datoprocesado[0],item2=datoprocesado[1],item3=datoprocesado[2])

            post = form.save(commit=False)

            post.save()

        else:
            datocrudo=form.data["ultrafile"]
            '''
            f = StringIO(datocrudo)
            reader = list(csv.reader(f, delimiter=','))
            print("hola")
            for row in reader:
                print('\t'.join(row))
                print("siguiente linea")
            #print(reader)
            '''
            datoprocesado=datocrudo.split(",;,")
            #print("datoprocesado1:")
            #print(datoprocesado)
            #print("Largo: " + str(len(datoprocesado)))
            for i in range (len(datoprocesado)):
                datoprocesado[i]=datoprocesado[i].split(",")

            #print("datoprocesado2:")
            #print(datoprocesado)
            '''
            print(datoprocesado[0])
            print("otro metodo:")

            print ((str.strip, s_inner.split(',')) for s_inner in datocrudo.split(";"))
            print("tercer metodo:")
            x = csv.reader(datocrudo)
            list(x)
            '''
            ####### identifica las columnas y crea los objetos que me interesanself.
            colFecha = None#Esta la tengo que pasar a datetime
            colTransindex = None
            colHorini = None#Esta la tengo que pasar a datetime
            colHorfin = None#Esta la tengo que pasar a datetime






            for i in range(len(datoprocesado[0])):
                if datoprocesado[0][i] == "CREACION_PROGRAMACION":
                    #print("columna creación en: " + str(i))
                    colFecha = i

                if datoprocesado[0][i] == "TRANSACTIONINDEX":
                    #print("columna transaction en: " + str(i))
                    colTransindex = i

                if datoprocesado[0][i] == "Horizonteini":
                    #print("columna Horizonteini en: " + str(i))
                    colHorini = i

                if datoprocesado[0][i] == "Horizontefin":
                    #print("columna Horizontefin en: " + str(i))
                    colHorfin= i

            '''
            #ejemplo

            obj, created = Person.objects.get_or_create(
                first_name='John',
                last_name='Lennon',
                defaults={'birthday': date(1940, 10, 9)},
            )
            '''

            fecha_programa_datetime=datetime.strptime(datoprocesado[1][colFecha], "%d-%m-%Y %H:%M")
            fecha_programa_horini=fecha_programa_datetime.replace(hour=0, minute=0, second=0, microsecond=0)
            fecha_programa_horfin=fecha_programa_horini + timedelta(days=1)
            #print("fecha programa")
            #print(fecha_programa_datetime)

            OrdenProg.objects.get_or_create(fecha_programa=fecha_programa_datetime, transaction_index=datoprocesado[1][colTransindex], horizonteini=fecha_programa_horini, horizontefin=fecha_programa_horfin )

            #################
            #################
            ## guardo el objeto DetalleProg por cada fila de la tabla procesada

            colworkcenter = None
            colorderid = None
            coldateini = None
            coldatefin = None
            colqin = None
            colidprev = None
            colidpost = None
            coldatefinajust = None
            colturno = None
            colanchoplaca= None
            collargoplaca = None
            colnumberout= None


            #print("Largo: " + str(len(datoprocesado[0])))
            for j in range(1,len(datoprocesado[0])):
                if datoprocesado[0][j] == "WORKCENTERID":
                    #print("columna colworkcenter en: " + str(j))
                    colworkcenter = j
                if datoprocesado[0][j] == "ORDERID":
                    #print("columna colorderid en: " + str(j))
                    colorderid = j
                if datoprocesado[0][j] == "SETUPSTARTDATE":
                    #print("columna dateini en: " + str(j))
                    coldateini = j
                if datoprocesado[0][j] == "RUNENDDATE":
                    #print("columna datefin en: " + str(j))
                    coldatefin = j
                if datoprocesado[0][j] == "QUANTITYIN":
                    #print("columna qin en: " + str(j))
                    colqin = j
                if datoprocesado[0][j] == "Fecha Termino Ajustada":
                    #print("columna datefinajust en: " + str(j))
                    coldatefinajust = j
                if datoprocesado[0][j] == "Turno":
                    #print("columna turno en: " + str(j))
                    colturno = j
                if datoprocesado[0][j] == "INPUTSHEETWIDTH":
                    #print("columna turno en: " + str(j))
                    colanchoplaca = j
                if datoprocesado[0][j] == "INPUTSHEETLENGTH":
                    #print("columna turno en: " + str(j))
                    collargoplaca = j
                if datoprocesado[0][j] == "NUMBEROUT":
                    #print("columna turno en: " + str(j))
                    colnumberout = j
            #print("Completado!!!!")

            for i in range(1,len(datoprocesado)):
                #try:
                    #print(OrdenProg.objects.all())



                    #print(datoprocesado[i])

                    datefinajustada_datetime = datetime.strptime(datoprocesado[i][coldatefinajust], "%d-%m-%Y")#"%d-%m-%Y %H:%M"
                    dateini_datetime = datetime.strptime(datoprocesado[i][coldateini], "%d-%m-%Y %H:%M")#"%d-%m-%Y %H:%M"
                    datefin_datetime = datetime.strptime(datoprocesado[i][coldatefin], "%d-%m-%Y %H:%M")#"%d-%m-%Y %H:%M"
                    cantQin=int(datoprocesado[i][colqin])
                    print("datefinajustada: " + str(datefinajustada_datetime))
                    print(OrdenProg.objects.filter(fecha_programa=fecha_programa_datetime, transaction_index=datoprocesado[1][colTransindex])[0])
                    print(datoprocesado[i][colworkcenter])
                    print(datoprocesado[i][colorderid])
                    print(datoprocesado[i][coldateini])
                    print(datoprocesado[i][coldatefin])
                    print(datoprocesado[i][colqin])
                    print(fecha_programa_datetime)
                    print(datoprocesado[i][colturno])
                    print(datoprocesado[i][colorderid])


                    DetalleProg.objects.get_or_create(programma=OrdenProg.objects.filter(fecha_programa=fecha_programa_datetime, transaction_index=datoprocesado[1][colTransindex])[0]  ,workcenter=datoprocesado[i][colworkcenter],orderId=datoprocesado[i][colorderid], dateini=dateini_datetime, datefin=datefin_datetime,qIn=cantQin, datefinajustada= datefinajustada_datetime , turno=datoprocesado[i][colturno], orderIdPrev=datoprocesado[i][colorderid], orderIdPost=datoprocesado[i][colorderid], anchoplaca=datoprocesado[i][colanchoplaca], largoplaca=datoprocesado[i][collargoplaca], numberout=datoprocesado[i][colnumberout])


                #except:

                    #print("hola")



            #PruebaTabla.objects.create(item1="form no valido11 !!", item2=datoprocesado[0], item3=datoprocesado,)



    else:
        form = PruebaModForm()

    return render(request, template_name, {'form':form})




def prueba_ultimate(request):
    pruebamods = PruebaMod.objects.all()
    posts= "HOLA Q ASE"
    template_name = 'blog/pruebaultimate.html'
    redirec_field_name = 'blog/pruebaultimate.html'

    if request.method == "POST":
        form = PruebaModForm(request.POST)
        if form.is_valid():

            '''
            ##Ejemplo de cómo crear un objeto de otra clase al llenar el form (para hacer el parse el ultimatefile y generar filas de info productiva)
            ##Se llena form de producto y si la categoría no existe, se crea el objeto de una categoría nueva (model Category)
            if form.is_valid():
                c = form.cleaned_data["category"]
                category = Category.objects.filter(name=c).first()
                if not category:
                    category = Category.objects.create(name=c)
            product = form.save(commit=False)
            product.category = category
            product.save()

            '''
            datocrudo=form.cleaned_data["ultrafile"]
            datoprocesado=datocrudo.split(",")

            PruebaTabla.objects.create(item1=datoprocesado[0],item2=datoprocesado[1],item3=datoprocesado[2])




            post = form.save(commit=False)
            #post.author = request.user
            #post.published_date = timezone.now()
            post.save()
            #return redirect('post_detail', pk=post.pk)
    else:
        form = PruebaModForm()



    return render(request, template_name, {'pruebamods':pruebamods, 'posts':posts, 'form':form})



def create_book_with_authors(request):
    template_name = 'blog/create_with_author.html'
    if request.method == 'GET':
        bookform = BookModelForm(request.GET or None)
        formset = AuthorFormset(queryset=Author.objects.none())
    elif request.method == 'POST':
        bookform = BookModelForm(request.POST)
        formset = AuthorFormset(request.POST)
        if bookform.is_valid() and formset.is_valid():
            # first save this book, as its reference will be used in `Author`
            book = bookform.save()
            for form in formset:
                # so that `book` instance can be attached.
                author = form.save(commit=False)
                author.book = book
                author.save()
            return redirect('blog:book_list')
    return render(request, template_name, {
        'bookform': bookform,
        'formset': formset,
    })

def create_book_model_form(request):
    template_name = 'blog/create_normal.html'
    heading_message = 'Model Formset Demo'
    if request.method == 'GET':
        # we don't want to display the already saved model instances
        formset = BookModelFormset(queryset=Book.objects.none())
    elif request.method == 'POST':
        formset = BookModelFormset(request.POST)
        if formset.is_valid():
            for form in formset:
                # only save if name is present
                if form.cleaned_data.get('name'):
                    form.save()
            return redirect('book_list')
    return render(request, template_name, {
        'formset': formset,
        'heading': heading_message,
    })




#Para el ejemplo de dynamic form creation
def create_book_normal(request):
    template_name = 'blog/create_normal.html'
    heading_message = 'Formset Demo'
    if request.method == 'GET':
        formset = BookFormset(request.GET or None)
    elif request.method == 'POST':
        formset = BookFormset(request.POST)
        if formset.is_valid():
            for form in formset:
                # extract name from each form and save
                name = form.cleaned_data.get('name')
                # save book instance
                if name:
                    Book(name=name).save()
            # once all books are saved, redirect to book list view
            return redirect('book_normal')
            #return redirect('book_list')
    return render(request, template_name, {
        'formset': formset,
        'heading': heading_message,
    })










def listaprodid(request):
    lista = ProdID.objects.all() #Tb podría haber sido filter
    context= {'lista_a_mostrar': lista}
    return render(request, 'blog/prodid_list.html', context)


class ProdIDListView(ListView):
    login_url = '/login/'
    #redirec_field_name = 'blog/post_detail.html'



    model = ProdID


    #redirect_field_name = 'blog/prodid_list.html'


    def get_queryset(self):
        return ProdID.objects.filter(published_date__isnull=True).order_by('-published_date')#sql query


def CargaCSV2(request):
    login_url = '/login/'
    redirec_field_name = 'blog/post_detail.html'
    with open("prueba2.csv") as f:
        reader = csv.reader(f, delimiter=';')
        for row in reader:
            _, created = ProdID.objects.get_or_create(
                transactindex =row[0],
                plantid=row[1],
                workcenter=row[2],
                orderid=row[4],
                title=row[4]+row[2],
                #middle_name=row[2],
                )

    def get_queryset(self):
        return ProdID.objects.filter(published_date__isnull=True).order_by('-published_date')#sql query
            # creates a tuple of the new object or
            # current object and a boolean of if it was created
    lista = ProdID.objects.all() #Tb podría haber sido filter
    context= {'lista_a_mostrar': lista}
    return render(request, 'blog/prodid_list.html', context)





def CargaCSV1(request):
    login_url = '/login/'
    redirec_field_name = 'blog/post_detail.html'
    with open("prueba.csv") as f:
        reader = csv.reader(f, delimiter=';')
        for row in reader:
            _, created = appointment.objects.get_or_create(
                title=row[0],
                text=row[1],
                #middle_name=row[2],
                )

    def get_queryset(self):
        return appointment.objects.filter(published_date__isnull=True).order_by('-published_date')#sql query
            # creates a tuple of the new object or
            # current object and a boolean of if it was created
    return render(request, 'blog/base.html')

def ContactView(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            return render(request, 'contact.html', {'form': form})
    else:
        form = ContactForm()
    return render(request, 'contact.html', {'form': form})





def product(x,y):
    return x*y



class CreateOCIView(LoginRequiredMixin,CreateView):
    login_url = '/login/'
    #redirec_field_name = 'blog/post_detail.html'

    form_class = OCImportacionForm

    model = OCImportacion

class appointmentCreate(LoginRequiredMixin, CreateView):
    model = appointment
    form_class = AppointmentForm

    def get_initial(self):
      #patient = self.request.GET.get('patient')
      patient = "dsds"
      location = "Hola Q hace"
      producto= product(3,5)
      tomanomb = tomanombre()


      return {
        'Patient': patient,
        'Location': location,
        #'Duration':producto,
        'Clinician': tomanomb,
      }



class AboutView(TemplateView):
    template_name = 'about.html'


'''
class SolCambListView( ListView):

    model = SolCamb
    redirect_field_name = 'blog/post_list.html'

    def get_queryset(self):
        return SolCamb.objects.filter(published_date__isnull=True).order_by('-published_date')#sql query


class SolCambListView(TemplateView):
    template_name = 'solcamb_list.html'

    model = SolCamb

    def get_queryset(self):
        return SolCamb.objects.filter(published_date__lte=timezone.now()).order_by('-published_date')#sql query

    '''


class PostListView(ListView):
    model = Post


    def get_queryset(self):
        return Post.objects.filter(published_date__lte=datetime.now()).order_by('-published_date')#sql query

class PostDetailView(DetailView):
    model = Post

class CreatePostView(LoginRequiredMixin,CreateView):
    login_url = '/login/'
    redirec_field_name = 'blog/post_detail.html'

    form_class = PostForm

    model = Post

class PostUpdateView(LoginRequiredMixin, UpdateView):
    login_url = '/login/'
    redirec_field_name = 'blog/post_detail.html'

    form_class = PostForm

    model = Post

class PostDeleteView(LoginRequiredMixin, DeleteView):
    model = Post
    success_url = reverse_lazy('post_list')

class DraftListView(LoginRequiredMixin, ListView):
    login_url = '/login/'
    redirect_field_name = 'blog/post_list.html'
    model = Post

    def get_queryset(self):
        return Post.objects.filter(published_date__isnull=True).order_by('create_date')


########################
########################

@login_required
def add_comment_to_post(request, pk):
    post = get_object_or_404(Post, pk=pk)

    if request.method == 'POST':
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.post = post
            comment.save()
            return redirect('post_detail', pk=post.pk)
    else:
        form = CommentForm()
    return render(request, 'blog/comment_form.html', {'form':form})

@login_required
def comment_approve(request, pk):
    comment = get_object_or_404(Comment, pk=pk)
    comment.approve()
    return redirect('post_detail', pk=comment.post.pk)


@login_required
def comment_remove(request, pk):
    comment = get_object_or_404(Comment, pk=pk)
    post_pk = comment.post.pk
    comment.delete()
    return redirect('post_detail', pk=post_pk)



@login_required
def post_publish(request, pk):
    post = get_object_or_404(Post, pk=pk)
    post.publish()
    return redirect('post_detail', pk=pk)




'''
@login_required
def editprofile(request):
    try:
    userprofile = UserProfiles.objects.get(user=request.user)
    except UserProfiles.DoesNotExist:
       return render(request, 'profile_edit.html', {'form':UserProfileForm()})
    form = UserProfileForm(instance=userprofile)
    return render(request, 'profile_edit.html', {'form':form})
'''

# Create your views here.
